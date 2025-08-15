# -*- coding: utf-8 -*-
"""
aggregate_cmd_tables.py

用途:
    掃描 `Centimani DOC/CMD＿TABLE` 目錄下所有 .xlsx 指令表，
    讀取每個分頁的表頭與部分資料，整合為 AI 可讀 JSON，
    並基於彙整結果輸出 NOKIA360 的腳本範本 JSON 與速查表 Markdown。

輸出:
    - AI_Complete_Reference/tools/NOKIA360_CMD_Tables_Combined.json
    - AI_Complete_Reference/tools/NOKIA360_Script_Template.json
    - AI_Complete_Reference/tools/NOKIA360_Command_CheatSheet.md
"""
from pathlib import Path
from typing import Dict, List, Tuple
import json
import re
import datetime
from openpyxl import load_workbook

HEADER_CANDIDATES = {
    "command": {"command", "cmd", "命令", "指令", "cli", "diag"},
    "description": {"description", "desc", "用途", "說明"},
    "example": {"example", "範例", "示例", "sample"},
    "reply": {"reply", "response", "回覆", "回應"},
    "category": {"category", "類別", "模組"},
}


def normalize_header(h: str) -> str:
    hs = (h or "").strip()
    hs_l = hs.lower()
    for key, names in HEADER_CANDIDATES.items():
        for n in names:
            if n in hs_l:
                return key
    return hs


def to_jsonable(v):
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.isoformat()
    # openpyxl 可能給出其他型別，保守轉字串
    try:
        json.dumps(v)
        return v
    except Exception:
        return str(v)


def read_workbook(xlsx: Path, sample_rows: int = 50) -> Dict:
    wb = load_workbook(xlsx, data_only=True)
    wb_info: Dict = {"file": str(xlsx), "sheets": []}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 讀 header
        raw_headers: List[str] = []
        norm_headers: List[str] = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            text = str(val).strip() if val is not None else ""
            raw_headers.append(text)
            norm_headers.append(normalize_header(text))
        # 讀 rows
        rows: List[Dict] = []
        for r in range(2, min(ws.max_row, sample_rows) + 1):
            row: Dict[str, object] = {}
            nonempty = False
            for c, h in enumerate(norm_headers, start=1):
                if not h:
                    continue
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    nonempty = True
                row[h] = to_jsonable(v)
            if nonempty:
                rows.append(row)
        wb_info["sheets"].append({
            "name": sheet_name,
            "headers": raw_headers,
            "norm_headers": norm_headers,
            "samples": rows,
        })
    return wb_info


def collect_candidate_commands(combined: Dict) -> List[Dict]:
    seen = set()
    commands: List[Dict] = []
    for f in combined.get("files", []):
        for s in f.get("sheets", []):
            for row in s.get("samples", []):
                cmd = row.get("command") or row.get("Command")
                if not cmd:
                    continue
                key = str(cmd).strip()
                if key and key not in seen:
                    seen.add(key)
                    commands.append({
                        "command": key,
                        "description": row.get("description"),
                        "reply": row.get("reply"),
                        "example": row.get("example"),
                        "source_sheet": s.get("name"),
                    })
                if len(commands) >= 200:
                    return commands
    return commands


def write_outputs(root: Path, combined: Dict) -> Tuple[Path, Path, Path]:
    tools_dir = root / "AI_Complete_Reference" / "tools"
    tools_dir.mkdir(parents=True, exist_ok=True)

    combined_json = tools_dir / "NOKIA360_CMD_Tables_Combined.json"
    combined_json.write_text(json.dumps(combined, ensure_ascii=False, indent=2), encoding="utf-8")

    # 產出腳本範本 JSON
    commands = collect_candidate_commands(combined)
    script_template = {
        "project": "NOKIA360",
        "sources": [f.get("file") for f in combined.get("files", [])],
        "commands_catalog": commands,
        "script_blueprints": [
            {
                "name": "版本號碼檢查",
                "fields": {
                    "TestID": "<請填入 TestID，如 HAVU002>",
                    "Phase": 1,
                    "Title": "版本號碼檢查",
                    "Send": "UART1@ver",
                    "Reply": "@Version:",
                    "Ref": "version",
                    "Waiting": "3000",
                    "ExecMode": "0",
                    "Description": "讀取並檢查版本字串"
                }
            },
            {
                "name": "電壓測量",
                "fields": {
                    "TestID": "<請查表填入>",
                    "Phase": 2,
                    "Title": "電壓測量",
                    "Instrument": "DMM.MeasureV",
                    "Send": ":DMM.MeasureV,\"CH1\",\"DC\"",
                    "Ref": "voltage",
                    "Validation": "(voltage,%f,4.75,5.25)",
                    "Waiting": "3000",
                    "ExecMode": "0",
                    "Description": "量測 5V 軌電壓"
                }
            },
            {
                "name": "BT狀態檢查",
                "fields": {
                    "TestID": "<請查表填入>",
                    "Phase": 3,
                    "Title": "BT狀態檢查",
                    "Send": "UART1@bt status",
                    "Reply": "@BT:ON",
                    "Waiting": "3000",
                    "ExecMode": "0",
                    "Description": "檢查藍牙啟用狀態"
                }
            }
        ]
    }
    template_json = tools_dir / "NOKIA360_Script_Template.json"
    template_json.write_text(json.dumps(script_template, ensure_ascii=False, indent=2), encoding="utf-8")

    # 產出速查表 Markdown
    md = []
    md.append("# NOKIA360 指令與腳本速查表\n")
    md.append("## 來源指令表\n")
    for f in combined.get("files", []):
        md.append(f"- {f.get('file')}")
    md.append("\n## 指令欄位映射建議\n")
    md.append("- command: 指令/命令/CMD/CLI/DIAG\n- description: 用途/說明\n- reply: 回覆/Response\n- example: 範例\n")
    md.append("\n## 代表性指令樣本 (最多200)\n")
    for i, c in enumerate(commands[:50], start=1):
        md.append(f"{i}. `{c.get('command')}` - {c.get('description') or ''}")
    md.append("\n## 腳本欄位填寫範例（三測項）\n")
    md.append("- 版本號碼: TestID=查表, Send=UART1@ver, Reply=@Version:, Ref=version\n")
    md.append("- 電壓: TestID=查表, Instrument=DMM.MeasureV, Send=:DMM.MeasureV,\"CH1\",\"DC\", Validation=(voltage,%f,4.75,5.25)\n")
    md.append("- BT: TestID=查表, Send=UART1@bt status, Reply=@BT:ON\n")
    md.append("\n## TestID 查詢\n")
    md.append("- 依 `System.ini > TestItemCodeFile` 或專案根目錄的 `Test Item Code V*.xlsx` 查表\n")
    md_path = tools_dir / "NOKIA360_Command_CheatSheet.md"
    md_path.write_text("\n".join(md), encoding="utf-8")

    return combined_json, template_json, md_path


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    cmd_dir = root / "Centimani DOC" / "CMD＿TABLE"
    if not cmd_dir.exists():
        raise SystemExit(f"找不到 CMD＿TABLE 目錄: {cmd_dir}")

    files = sorted([p for p in cmd_dir.iterdir() if p.suffix.lower() == ".xlsx"])
    combined = {"dir": str(cmd_dir), "files": []}
    for f in files:
        try:
            info = read_workbook(f)
            combined["files"].append(info)
        except Exception as e:
            combined.setdefault("errors", []).append({"file": str(f), "error": str(e)})

    cjson, tjson, md = write_outputs(root, combined)
    print(f"已輸出: {cjson}")
    print(f"已輸出: {tjson}")
    print(f"已輸出: {md}")


if __name__ == "__main__":
    main() 