#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Centimani 資料整理工具
主要功能：
1. 分析 Centimani DOC 資料夾中的所有 Excel 檔案
2. 解析 INI 檔案設定
3. 生成整合的速查表
4. 創建範例資料
5. 輸出整理後的資料結構

作者：AI Assistant
版本：1.0.0
日期：2024-12-19
"""

import os
import pandas as pd
import json
import csv
from pathlib import Path
import logging
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import configparser
import re

class CentimaniDataOrganizer:
    """Centimani 資料整理器"""
    
    def __init__(self, base_path="Centimani DOC"):
        """初始化資料整理器"""
        self.base_path = Path(base_path)
        self.output_path = Path("organized_data")
        self.setup_logging()
        self.setup_output_directories()
        
        # 資料儲存
        self.excel_summaries = []
        self.ini_settings = {}
        self.test_flows = []
        self.command_references = []
        
    def setup_logging(self):
        """設定日誌記錄"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('centimani_organizer.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def setup_output_directories(self):
        """建立輸出目錄"""
        directories = [
            self.output_path,
            self.output_path / "excel_summaries",
            self.output_path / "ini_settings",
            self.output_path / "test_flows",
            self.output_path / "command_references",
            self.output_path / "quick_reference"
        ]
        
        for directory in directories:
            directory.mkdir(parents=True, exist_ok=True)
            
    def analyze_excel_files(self):
        """分析所有 Excel 檔案"""
        self.logger.info("開始分析 Excel 檔案...")
        
        excel_files = []
        for root, dirs, files in os.walk(self.base_path):
            for file in files:
                if file.endswith(('.xlsx', '.xls')):
                    excel_files.append(Path(root) / file)
        
        self.logger.info(f"找到 {len(excel_files)} 個 Excel 檔案")
        
        for excel_file in excel_files:
            try:
                self.analyze_single_excel(excel_file)
            except Exception as e:
                self.logger.error(f"分析檔案 {excel_file} 時發生錯誤: {e}")
                
    def analyze_single_excel(self, file_path):
        """分析單一 Excel 檔案"""
        try:
            # 使用 openpyxl 讀取檔案
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            file_info = {
                'file_name': file_path.name,
                'file_path': str(file_path.relative_to(self.base_path)),
                'file_size': file_path.stat().st_size,
                'sheets': [],
                'total_rows': 0,
                'total_columns': 0,
                'keywords': [],
                'test_items': []
            }
            
            # 分析每個工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info = self.analyze_sheet(sheet, sheet_name)
                file_info['sheets'].append(sheet_info)
                file_info['total_rows'] += sheet_info['rows']
                file_info['total_columns'] = max(file_info['total_columns'], sheet_info['columns'])
                
            # 提取關鍵字和測試項目
            file_info['keywords'] = self.extract_keywords(file_info['sheets'])
            file_info['test_items'] = self.extract_test_items(file_info['sheets'])
            
            self.excel_summaries.append(file_info)
            self.logger.info(f"成功分析: {file_path.name}")
            
        except Exception as e:
            self.logger.error(f"無法讀取 {file_path}: {e}")
            
    def analyze_sheet(self, sheet, sheet_name):
        """分析工作表"""
        sheet_info = {
            'name': sheet_name,
            'rows': sheet.max_row,
            'columns': sheet.max_column,
            'data_types': {},
            'sample_data': []
        }
        
        # 分析前幾行資料作為範例
        for row in range(1, min(6, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(11, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    row_data.append(str(cell_value))
                    # 記錄資料類型
                    data_type = type(cell_value).__name__
                    sheet_info['data_types'][data_type] = sheet_info['data_types'].get(data_type, 0) + 1
            if row_data:
                sheet_info['sample_data'].append(row_data)
                
        return sheet_info
        
    def extract_keywords(self, sheets):
        """提取關鍵字"""
        keywords = set()
        common_keywords = [
            'test', 'flow', 'command', 'parameter', 'setting', 'config',
            'kangaroo', 'valo360', 'pega', 'nokia', 'function', 'check',
            'verification', 'calibration', 'download', 'upgrade', 'usb',
            'wifi', 'bluetooth', 'cellular', 'ota', 'fw', 'firmware'
        ]
        
        for sheet in sheets:
            for row_data in sheet['sample_data']:
                for cell_value in row_data:
                    cell_lower = str(cell_value).lower()
                    for keyword in common_keywords:
                        if keyword in cell_lower:
                            keywords.add(keyword)
                            
        return list(keywords)
        
    def extract_test_items(self, sheets):
        """提取測試項目"""
        test_items = []
        test_patterns = [
            r'test\s*item\s*code',
            r'test\s*flow',
            r'verification',
            r'calibration',
            r'function\s*check'
        ]
        
        for sheet in sheets:
            for row_data in sheet['sample_data']:
                for cell_value in row_data:
                    cell_str = str(cell_value).lower()
                    for pattern in test_patterns:
                        if re.search(pattern, cell_str):
                            test_items.append({
                                'sheet': sheet['name'],
                                'value': str(cell_value),
                                'pattern': pattern
                            })
                            break
                            
        return test_items
        
    def analyze_ini_files(self):
        """分析 INI 檔案"""
        self.logger.info("開始分析 INI 檔案...")
        
        ini_files = list(self.base_path.glob("*.ini"))
        self.logger.info(f"找到 {len(ini_files)} 個 INI 檔案")
        
        for ini_file in ini_files:
            try:
                self.analyze_single_ini(ini_file)
            except Exception as e:
                self.logger.error(f"分析 INI 檔案 {ini_file} 時發生錯誤: {e}")
                
    def analyze_single_ini(self, file_path):
        """分析單一 INI 檔案"""
        try:
            config = configparser.ConfigParser()
            config.read(file_path, encoding='utf-8')
            
            ini_info = {
                'file_name': file_path.name,
                'file_path': str(file_path.relative_to(self.base_path)),
                'sections': [],
                'total_settings': 0
            }
            
            for section_name in config.sections():
                section_info = {
                    'name': section_name,
                    'settings': dict(config[section_name]),
                    'setting_count': len(config[section_name])
                }
                ini_info['sections'].append(section_info)
                ini_info['total_settings'] += section_info['setting_count']
                
            self.ini_settings[file_path.name] = ini_info
            self.logger.info(f"成功分析 INI 檔案: {file_path.name}")
            
        except Exception as e:
            self.logger.error(f"無法讀取 INI 檔案 {file_path}: {e}")
            
    def generate_quick_reference(self):
        """生成速查表"""
        self.logger.info("生成速查表...")
        
        # 1. Excel 檔案總覽
        excel_overview = {
            'total_files': len(self.excel_summaries),
            'total_size_mb': sum(f['file_size'] for f in self.excel_summaries) / (1024*1024),
            'file_categories': self.categorize_files(),
            'keywords_summary': self.summarize_keywords()
        }
        
        # 2. INI 設定總覽
        ini_overview = {
            'total_files': len(self.ini_settings),
            'total_settings': sum(info['total_settings'] for info in self.ini_settings.values()),
            'common_sections': self.find_common_sections()
        }
        
        # 3. 測試流程分類
        test_flow_categories = self.categorize_test_flows()
        
        # 4. 命令參考
        command_reference = self.generate_command_reference()
        
        # 儲存速查表
        quick_ref = {
            'generated_date': datetime.now().isoformat(),
            'excel_overview': excel_overview,
            'ini_overview': ini_overview,
            'test_flow_categories': test_flow_categories,
            'command_reference': command_reference
        }
        
        # 儲存為 JSON
        with open(self.output_path / "quick_reference" / "quick_reference.json", 'w', encoding='utf-8') as f:
            json.dump(quick_ref, f, ensure_ascii=False, indent=2)
            
        # 儲存為 CSV
        self.save_quick_reference_csv(quick_ref)
        
        self.logger.info("速查表生成完成")
        
    def categorize_files(self):
        """分類檔案"""
        categories = {
            'KANGAROO': [],
            'VALO360': [],
            'Centimani': [],
            'Test_Items': [],
            'Other': []
        }
        
        for file_info in self.excel_summaries:
            file_name = file_info['file_name'].lower()
            if 'kangaroo' in file_name:
                categories['KANGAROO'].append(file_info['file_name'])
            elif 'valo360' in file_name:
                categories['VALO360'].append(file_info['file_name'])
            elif 'centimani' in file_name:
                categories['Centimani'].append(file_info['file_name'])
            elif 'test item' in file_name or 'testitem' in file_name:
                categories['Test_Items'].append(file_info['file_name'])
            else:
                categories['Other'].append(file_info['file_name'])
                
        return categories
        
    def summarize_keywords(self):
        """總結關鍵字"""
        keyword_count = {}
        for file_info in self.excel_summaries:
            for keyword in file_info['keywords']:
                keyword_count[keyword] = keyword_count.get(keyword, 0) + 1
                
        return sorted(keyword_count.items(), key=lambda x: x[1], reverse=True)
        
    def find_common_sections(self):
        """找出常見的 INI 區段"""
        section_count = {}
        for ini_info in self.ini_settings.values():
            for section in ini_info['sections']:
                section_name = section['name']
                section_count[section_name] = section_count.get(section_name, 0) + 1
                
        return sorted(section_count.items(), key=lambda x: x[1], reverse=True)
        
    def categorize_test_flows(self):
        """分類測試流程"""
        categories = {
            'KANGAROO': {
                'Mainboard': [],
                'OQC': [],
                'Cellular': [],
                'WIFI_BT': [],
                'OTA': [],
                'MultiDUT': []
            },
            'VALO360': {
                'ER2': [],
                'POE': [],
                'IMU': [],
                'Function': [],
                'DMIC': []
            }
        }
        
        for file_info in self.excel_summaries:
            file_name = file_info['file_name'].lower()
            if 'kangaroo' in file_name:
                if 'mainboard' in file_name:
                    categories['KANGAROO']['Mainboard'].append(file_info['file_name'])
                elif 'oqc' in file_name:
                    categories['KANGAROO']['OQC'].append(file_info['file_name'])
                elif 'cellular' in file_name:
                    categories['KANGAROO']['Cellular'].append(file_info['file_name'])
                elif 'wifi' in file_name or 'bt' in file_name:
                    categories['KANGAROO']['WIFI_BT'].append(file_info['file_name'])
                elif 'ota' in file_name:
                    categories['KANGAROO']['OTA'].append(file_info['file_name'])
                elif 'multidut' in file_name:
                    categories['KANGAROO']['MultiDUT'].append(file_info['file_name'])
            elif 'valo360' in file_name:
                if 'er2' in file_name:
                    categories['VALO360']['ER2'].append(file_info['file_name'])
                elif 'poe' in file_name:
                    categories['VALO360']['POE'].append(file_info['file_name'])
                elif 'imu' in file_name:
                    categories['VALO360']['IMU'].append(file_info['file_name'])
                elif 'function' in file_name:
                    categories['VALO360']['Function'].append(file_info['file_name'])
                elif 'dmic' in file_name:
                    categories['VALO360']['DMIC'].append(file_info['file_name'])
                    
        return categories
        
    def generate_command_reference(self):
        """生成命令參考"""
        commands = []
        
        # 從 INI 檔案提取命令
        for ini_name, ini_info in self.ini_settings.items():
            for section in ini_info['sections']:
                for key, value in section['settings'].items():
                    if any(cmd_word in key.lower() for cmd_word in ['command', 'cmd', 'script']):
                        commands.append({
                            'source': ini_name,
                            'section': section['name'],
                            'command_key': key,
                            'command_value': value
                        })
                        
        return commands
        
    def save_quick_reference_csv(self, quick_ref):
        """儲存速查表為 CSV 格式"""
        # Excel 檔案總覽
        excel_data = []
        for category, files in quick_ref['excel_overview']['file_categories'].items():
            for file_name in files:
                excel_data.append({
                    'Category': category,
                    'File_Name': file_name,
                    'Type': 'Excel'
                })
                
        df_excel = pd.DataFrame(excel_data)
        df_excel.to_csv(self.output_path / "quick_reference" / "excel_files_overview.csv", 
                        index=False, encoding='utf-8-sig')
        
        # INI 設定總覽
        ini_data = []
        for ini_name, ini_info in self.ini_settings.items():
            for section in ini_info['sections']:
                for key, value in section['settings'].items():
                    ini_data.append({
                        'INI_File': ini_name,
                        'Section': section['name'],
                        'Setting_Key': key,
                        'Setting_Value': value
                    })
                    
        df_ini = pd.DataFrame(ini_data)
        df_ini.to_csv(self.output_path / "quick_reference" / "ini_settings_overview.csv", 
                      index=False, encoding='utf-8-sig')
        
        # 測試流程分類
        test_flow_data = []
        for platform, categories in quick_ref['test_flow_categories'].items():
            for category, files in categories.items():
                for file_name in files:
                    test_flow_data.append({
                        'Platform': platform,
                        'Category': category,
                        'File_Name': file_name
                    })
                    
        df_test_flow = pd.DataFrame(test_flow_data)
        df_test_flow.to_csv(self.output_path / "quick_reference" / "test_flow_categories.csv", 
                           index=False, encoding='utf-8-sig')
        
    def generate_examples(self):
        """生成範例資料"""
        self.logger.info("生成範例資料...")
        
        # 1. Excel 檔案分析範例
        excel_examples = []
        for file_info in self.excel_summaries[:5]:  # 取前5個檔案作為範例
            example = {
                'file_name': file_info['file_name'],
                'structure': {
                    'sheets': [sheet['name'] for sheet in file_info['sheets']],
                    'total_rows': file_info['total_rows'],
                    'total_columns': file_info['total_columns']
                },
                'sample_data': file_info['sheets'][0]['sample_data'][:3] if file_info['sheets'] else [],
                'keywords': file_info['keywords'][:5]
            }
            excel_examples.append(example)
            
        # 2. INI 設定範例
        ini_examples = []
        for ini_name, ini_info in list(self.ini_settings.items())[:3]:
            example = {
                'file_name': ini_name,
                'sections': [section['name'] for section in ini_info['sections']],
                'sample_section': ini_info['sections'][0] if ini_info['sections'] else {}
            }
            ini_examples.append(example)
            
        # 儲存範例
        examples = {
            'excel_examples': excel_examples,
            'ini_examples': ini_examples,
            'generated_date': datetime.now().isoformat()
        }
        
        with open(self.output_path / "examples.json", 'w', encoding='utf-8') as f:
            json.dump(examples, f, ensure_ascii=False, indent=2)
            
        self.logger.info("範例資料生成完成")
        
    def run_analysis(self):
        """執行完整分析"""
        self.logger.info("開始 Centimani 資料整理分析...")
        
        # 1. 分析 Excel 檔案
        self.analyze_excel_files()
        
        # 2. 分析 INI 檔案
        self.analyze_ini_files()
        
        # 3. 生成速查表
        self.generate_quick_reference()
        
        # 4. 生成範例資料
        self.generate_examples()
        
        # 5. 生成總結報告
        self.generate_summary_report()
        
        self.logger.info("資料整理分析完成！")
        
    def generate_summary_report(self):
        """生成總結報告"""
        self.logger.info("生成總結報告...")
        
        report = f"""
# Centimani 資料整理報告

## 分析時間
{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## 檔案統計
- Excel 檔案總數: {len(self.excel_summaries)}
- INI 檔案總數: {len(self.ini_settings)}
- 總資料大小: {sum(f['file_size'] for f in self.excel_summaries) / (1024*1024):.2f} MB

## 主要分類
### KANGAROO 平台
- 主機板測試流程
- OQC 測試流程  
- 蜂窩網路驗證
- WiFi & Bluetooth 校準驗證
- OTA 韌體升級
- 多設備測試

### VALO360 平台
- ER2 功能檢查
- POE 測試流程
- IMU 測試流程
- 功能驗證測試
- DMIC 測試流程

## 輸出檔案
- 速查表: organized_data/quick_reference/
- Excel 分析: organized_data/excel_summaries/
- INI 設定: organized_data/ini_settings/
- 範例資料: organized_data/examples.json

## 使用建議
1. 使用 quick_reference 資料夾中的 CSV 檔案進行快速查詢
2. 參考 examples.json 了解資料結構
3. 根據分類找到相關的測試流程檔案
4. 使用 INI 設定檔案作為系統配置參考
"""
        
        with open(self.output_path / "summary_report.md", 'w', encoding='utf-8') as f:
            f.write(report)
            
        self.logger.info("總結報告生成完成")

def main():
    """主程式"""
    print("Centimani 資料整理工具")
    print("=" * 50)
    
    # 檢查資料夾是否存在
    if not Path("Centimani DOC").exists():
        print("錯誤: 找不到 'Centimani DOC' 資料夾")
        print("請確保資料夾存在於當前目錄中")
        return
        
    # 創建整理器並執行分析
    organizer = CentimaniDataOrganizer()
    organizer.run_analysis()
    
    print("\n資料整理完成！")
    print(f"輸出目錄: {organizer.output_path}")
    print("\n主要輸出檔案:")
    print("- quick_reference/ - 速查表")
    print("- examples.json - 範例資料")
    print("- summary_report.md - 總結報告")
    
if __name__ == "__main__":
    main() 