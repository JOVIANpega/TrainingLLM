#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Centimani格式測試腳本

本腳本按照Centimani測試腳本的標準格式創建，包含：
1. Properties工作表 - 版本號碼和基本資訊
2. Switch工作表 - 開關配置
3. Instrument工作表 - 儀器配置
4. DUTs工作表 - 被測設備配置
5. 測試項目工作表 - 具體的測試流程

符合Centimani腳本的標準結構和格式要求

作者: AI Assistant
版本: 1.0.0
更新日期: 2024-12-19
"""

import os
import json
import time
import random
from datetime import datetime
from typing import Dict, List, Tuple, Any
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class CentimaniTestScript:
    """Centimani格式測試腳本生成器"""
    
    def __init__(self):
        self.script_data = {
            "properties": {},
            "switches": [],
            "instruments": [],
            "duts": [],
            "test_items": []
        }
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
    def set_properties(self, name: str, version: str, creator: str = "AI Assistant"):
        """設置腳本屬性"""
        self.script_data["properties"] = {
            "Name": name,
            "Version": version,
            "Creator": creator,
            "FailCount": "0",
            "CreateDate": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Description": "測試量測腳本 - 版本管理、電壓量測、藍牙量測",
            "TestType": "Measurement",
            "Category": "TestFlow",
            "Priority": "High",
            "Status": "Active"
        }
    
    def add_switch(self, name: str, control_dll: str, settings: str, version: str = "1.0.0.0", loss: str = "0", desc: str = ""):
        """添加開關配置"""
        switch = {
            "Mount": "",
            "Name": name,
            "ControlDLL": control_dll,
            "Settings": settings,
            "Version": version,
            "Loss": loss,
            "Desc": desc
        }
        self.script_data["switches"].append(switch)
    
    def add_instrument(self, mount: str, name: str, sub: str, control_dll: str, settings: str, functions: str, calibration: str = "", status: str = "Active"):
        """添加儀器配置"""
        instrument = {
            "Mount": mount,
            "Name": name,
            "Sub": sub,
            "ControlDLL": control_dll,
            "Settings": settings,
            "Functions": functions,
            "Calibration": calibration,
            "Status": status
        }
        self.script_data["instruments"].append(instrument)
    
    def add_dut(self, mount: str, name: str, type_name: str, settings: str, description: str = ""):
        """添加被測設備配置"""
        dut = {
            "Mount": mount,
            "Name": name,
            "Type": type_name,
            "Settings": settings,
            "Description": description,
            "Status": "Active",
            "TestCount": "0",
            "PassCount": "0",
            "FailCount": "0"
        }
        self.script_data["duts"].append(dut)
    
    def add_test_item(self, step: str, test_name: str, instrument: str, dut: str, parameters: str, expected: str, timeout: str = "30", description: str = ""):
        """添加測試項目"""
        test_item = {
            "Step": step,
            "TestName": test_name,
            "Instrument": instrument,
            "DUT": dut,
            "Parameters": parameters,
            "Expected": expected,
            "Timeout": timeout,
            "Description": description,
            "Status": "Pending",
            "Result": "",
            "Value": "",
            "Unit": "",
            "Pass/Fail": "",
            "Comments": ""
        }
        self.script_data["test_items"].append(test_item)
    
    def generate_voltage_measurement_tests(self):
        """生成電壓量測測試項目"""
        # 電壓量測測試項目
        voltage_tests = [
            {
                "step": "1",
                "test_name": "Voltage_Measurement_Channel1",
                "instrument": "VoltageMeter",
                "dut": "DUT1",
                "parameters": "Channel=1, Range=5V, Accuracy=0.01V",
                "expected": "3.3V ±0.1V",
                "description": "量測通道1電壓值"
            },
            {
                "step": "2",
                "test_name": "Voltage_Measurement_Channel2",
                "instrument": "VoltageMeter",
                "dut": "DUT1",
                "parameters": "Channel=2, Range=5V, Accuracy=0.01V",
                "expected": "3.3V ±0.1V",
                "description": "量測通道2電壓值"
            },
            {
                "step": "3",
                "test_name": "Voltage_Measurement_Channel3",
                "instrument": "VoltageMeter",
                "dut": "DUT1",
                "parameters": "Channel=3, Range=5V, Accuracy=0.01V",
                "expected": "3.3V ±0.1V",
                "description": "量測通道3電壓值"
            },
            {
                "step": "4",
                "test_name": "Voltage_Stability_Test",
                "instrument": "VoltageMeter",
                "dut": "DUT1",
                "parameters": "Duration=60s, Sampling=1Hz",
                "expected": "Stable ±0.05V",
                "description": "電壓穩定性測試"
            }
        ]
        
        for test in voltage_tests:
            self.add_test_item(**test)
    
    def generate_bluetooth_measurement_tests(self):
        """生成藍牙量測測試項目"""
        # 藍牙量測測試項目
        bluetooth_tests = [
            {
                "step": "5",
                "test_name": "Bluetooth_Device_Scan",
                "instrument": "BluetoothTester",
                "dut": "DUT1",
                "parameters": "ScanDuration=10s, RSSI_Threshold=-80dBm",
                "expected": "Device_Count >= 1",
                "description": "藍牙設備掃描測試"
            },
            {
                "step": "6",
                "test_name": "Bluetooth_Connection_Test",
                "instrument": "BluetoothTester",
                "dut": "DUT1",
                "parameters": "Device=TestDevice, Timeout=30s",
                "expected": "Connection_Success",
                "description": "藍牙連接測試"
            },
            {
                "step": "7",
                "test_name": "Bluetooth_RSSI_Measurement",
                "instrument": "BluetoothTester",
                "dut": "DUT1",
                "parameters": "Device=TestDevice, Duration=10s",
                "expected": "RSSI >= -70dBm",
                "description": "藍牙信號強度量測"
            },
            {
                "step": "8",
                "test_name": "Bluetooth_Profile_Test",
                "instrument": "BluetoothTester",
                "dut": "DUT1",
                "parameters": "Profiles=HFP,A2DP,AVRCP,GATT",
                "expected": "All_Profiles_Supported",
                "description": "藍牙協議支援測試"
            }
        ]
        
        for test in bluetooth_tests:
            self.add_test_item(**test)
    
    def generate_version_management_tests(self):
        """生成版本管理測試項目"""
        # 版本管理測試項目
        version_tests = [
            {
                "step": "9",
                "test_name": "Version_Check",
                "instrument": "SystemInfo",
                "dut": "DUT1",
                "parameters": "CheckType=SoftwareVersion",
                "expected": "Version_Valid",
                "description": "軟體版本檢查"
            },
            {
                "step": "10",
                "test_name": "Build_Number_Verification",
                "instrument": "SystemInfo",
                "dut": "DUT1",
                "parameters": "CheckType=BuildNumber",
                "expected": "Build_Number_Valid",
                "description": "建置編號驗證"
            },
            {
                "step": "11",
                "test_name": "Version_Compatibility_Test",
                "instrument": "SystemInfo",
                "dut": "DUT1",
                "parameters": "CheckType=Compatibility",
                "expected": "Compatible",
                "description": "版本相容性測試"
            }
        ]
        
        for test in version_tests:
            self.add_test_item(**test)
    
    def create_excel_workbook(self) -> openpyxl.Workbook:
        """創建Excel工作簿"""
        wb = openpyxl.Workbook()
        
        # 移除預設工作表
        wb.remove(wb.active)
        
        # 創建Properties工作表
        self._create_properties_sheet(wb)
        
        # 創建Switch工作表
        self._create_switch_sheet(wb)
        
        # 創建Instrument工作表
        self._create_instrument_sheet(wb)
        
        # 創建DUTs工作表
        self._create_duts_sheet(wb)
        
        # 創建測試項目工作表
        self._create_test_items_sheet(wb)
        
        return wb
    
    def _create_properties_sheet(self, wb: openpyxl.Workbook):
        """創建Properties工作表"""
        ws = wb.create_sheet("Properties")
        
        # 設置標題樣式
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 設置標題
        ws['A1'] = "Name"
        ws['B1'] = "Value"
        ws['A1'].font = title_font
        ws['B1'].font = title_font
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        
        # 添加屬性數據
        row = 2
        for key, value in self.script_data["properties"].items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_switch_sheet(self, wb: openpyxl.Workbook):
        """創建Switch工作表"""
        ws = wb.create_sheet("Switch")
        
        # 設置標題樣式
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 設置標題行
        headers = ["Mount", "Name", "ControlDLL", "Settings", "Version", "Loss", "Desc"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = title_font
            cell.fill = header_fill
        
        # 添加開關數據
        for row, switch in enumerate(self.script_data["switches"], 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=switch.get(key, ""))
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_instrument_sheet(self, wb: openpyxl.Workbook):
        """創建Instrument工作表"""
        ws = wb.create_sheet("Instrument")
        
        # 設置標題樣式
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 設置標題行
        headers = ["Mount", "Name", "Sub", "ControlDLL", "Settings", "Functions", "Calibration", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = title_font
            cell.fill = header_fill
        
        # 添加儀器數據
        for row, instrument in enumerate(self.script_data["instruments"], 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=instrument.get(key, ""))
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_duts_sheet(self, wb: openpyxl.Workbook):
        """創建DUTs工作表"""
        ws = wb.create_sheet("DUTs")
        
        # 設置標題樣式
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 設置標題行
        headers = ["Mount", "Name", "Type", "Settings", "Description", "Status", "TestCount", "PassCount", "FailCount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = title_font
            cell.fill = header_fill
        
        # 添加DUT數據
        for row, dut in enumerate(self.script_data["duts"], 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=dut.get(key, ""))
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_test_items_sheet(self, wb: openpyxl.Workbook):
        """創建測試項目工作表"""
        ws = wb.create_sheet("Test_Items")
        
        # 設置標題樣式
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 設置標題行
        headers = [
            "Step", "TestName", "Instrument", "DUT", "Parameters", "Expected", 
            "Timeout", "Description", "Status", "Result", "Value", "Unit", 
            "Pass/Fail", "Comments"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = title_font
            cell.fill = header_fill
        
        # 添加測試項目數據
        for row, test_item in enumerate(self.script_data["test_items"], 2):
            for col, key in enumerate(headers, 1):
                # 將字典鍵轉換為對應的測試項目鍵
                if key == "TestName":
                    value = test_item.get("TestName", "")
                elif key == "Pass/Fail":
                    value = test_item.get("Pass/Fail", "")
                else:
                    value = test_item.get(key.lower().replace("/", "").replace(" ", ""), "")
                
                ws.cell(row=row, column=col, value=value)
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def save_excel_file(self, filename: str = None) -> str:
        """保存Excel檔案"""
        if filename is None:
            filename = f"Centimani_TestScript_{self.timestamp}.xlsx"
        
        # 創建工作簿
        wb = self.create_excel_workbook()
        
        # 保存檔案
        wb.save(filename)
        wb.close()
        
        return filename
    
    def generate_script(self, script_name: str = "TestMeasurementScript") -> str:
        """生成完整的測試腳本"""
        # 設置腳本屬性
        self.set_properties(
            name=script_name,
            version=f"1.0.0-{self.timestamp}",
            creator="AI Assistant"
        )
        
        # 添加開關配置
        self.add_switch(
            name="VoltageSwitch",
            control_dll="VoltageSwitch.dll",
            settings="GPIB,GPIB0::2::INSTR,1,1",
            desc="電壓量測開關"
        )
        
        self.add_switch(
            name="BluetoothSwitch",
            control_dll="BluetoothSwitch.dll",
            settings="USB,USB0::0x1234::0x5678::INSTR",
            desc="藍牙測試開關"
        )
        
        # 添加儀器配置
        self.add_instrument(
            mount="1",
            name="VoltageMeter",
            sub="1",
            control_dll="VoltageMeter.dll",
            settings="GPIB,GPIB0::3::INSTR",
            functions="DC_Voltage,AC_Voltage,Resistance",
            calibration="2024-12-19"
        )
        
        self.add_instrument(
            mount="2",
            name="BluetoothTester",
            sub="1",
            control_dll="BluetoothTester.dll",
            settings="USB,USB0::0x1234::0x5678::INSTR",
            functions="Device_Scan,Connection_Test,RSSI_Measurement",
            calibration="2024-12-19"
        )
        
        self.add_instrument(
            mount="3",
            name="SystemInfo",
            sub="1",
            control_dll="SystemInfo.dll",
            settings="Local",
            functions="Version_Check,Build_Verification,Compatibility_Test",
            calibration="N/A"
        )
        
        # 添加被測設備
        self.add_dut(
            mount="1",
            name="DUT1",
            type_name="TestBoard",
            settings="Port=COM1,BaudRate=115200",
            description="測試電路板"
        )
        
        # 生成測試項目
        self.generate_voltage_measurement_tests()
        self.generate_bluetooth_measurement_tests()
        self.generate_version_management_tests()
        
        # 保存Excel檔案
        filename = self.save_excel_file()
        
        return filename


def main():
    """主函數"""
    print("=" * 60)
    print("Centimani格式測試腳本生成器")
    print("=" * 60)
    
    try:
        # 創建腳本生成器
        script_generator = CentimaniTestScript()
        
        # 生成測試腳本
        script_name = "TestMeasurementScript_Voltage_BT_Version"
        filename = script_generator.generate_script(script_name)
        
        print(f"腳本生成完成！")
        print(f"檔案名稱: {filename}")
        print(f"腳本名稱: {script_name}")
        print(f"版本: {script_generator.script_data['properties']['Version']}")
        print(f"創建者: {script_generator.script_data['properties']['Creator']}")
        print(f"測試項目數量: {len(script_generator.script_data['test_items'])}")
        print(f"儀器數量: {len(script_generator.script_data['instruments'])}")
        print(f"被測設備數量: {len(script_generator.script_data['duts'])}")
        
        print(f"\n腳本已保存到: {filename}")
        
    except Exception as e:
        print(f"腳本生成失敗: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main() 