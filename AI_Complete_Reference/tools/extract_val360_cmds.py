# -*- coding: utf-8 -*-
"""
extract_val360_cmds.py

用途:
    讀取 `Centimani DOC/CMD＿TABLE/VAL360_Diag_Command_List_Requirement_20250627.xlsx`
    輸出欄位結構、各分頁前若干列樣本，並保存為 JSON 摘要，供 AI 參考。
輸出:
    AI_Complete_Reference/tools/VAL360_CMD_TABLE_Summary.json
"""
from pathlib import Path
import json
from openpyxl import load_workbook


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    xlsx = root / "Centimani DOC" / "CMD＿TABLE" / "VAL360_Diag_Command_List_Requirement_20250627.xlsx"
    if not xlsx.exists():
        raise SystemExit(f"找不到檔案: {xlsx}")

    wb = load_workbook(xlsx, data_only=True)
    summary = {"file": str(xlsx), "sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val is None:
                val = ""
            headers.append(str(val).strip())
        rows = []
        for r in range(2, min(ws.max_row, 20) + 1):
            row = {}
            any_val = False
            for c, h in enumerate(headers, start=1):
                if not h:
                    continue
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    any_val = True
                row[h] = v
            if any_val:
                rows.append(row)
        summary["sheets"].append({"name": sheet_name, "headers": headers, "samples": rows})

    out_json = Path(__file__).resolve().parent / "VAL360_CMD_TABLE_Summary.json"
    out_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"已輸出: {out_json}")


if __name__ == "__main__":
    main() 