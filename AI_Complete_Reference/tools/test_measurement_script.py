#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
測試量測腳本

本腳本提供以下功能：
1. 版本號碼管理 - 自動生成和管理版本資訊
2. 量測電壓 - 模擬電壓量測並記錄數據
3. 量測BT - 模擬藍牙量測並記錄數據
4. 數據輸出 - 生成CSV和XLSX格式的測試報告

主要用途：
- 測試環境的數據收集和分析
- 生成標準化的測試報告
- 支援多種輸出格式
- 自動化測試流程

作者: AI Assistant
版本: 1.0.0
更新日期: 2024-12-19
"""

import os
import json
import time
import random
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('test_measurement.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class VersionManager:
    """版本號碼管理類"""
    
    def __init__(self, version_file: str = "version_info.json"):
        self.version_file = version_file
        self.version_info = self.load_version_info()
    
    def load_version_info(self) -> Dict[str, Any]:
        """載入版本資訊"""
        default_version = {
            "major": 1,
            "minor": 0,
            "patch": 0,
            "build": 1,
            "release_date": datetime.now().isoformat(),
            "description": "初始版本",
            "changes": ["初始版本發布"]
        }
        
        try:
            if os.path.exists(self.version_file):
                with open(self.version_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                self.save_version_info(default_version)
                return default_version
        except Exception as e:
            logger.error(f"載入版本資訊失敗: {e}")
            return default_version
    
    def save_version_info(self, version_info: Dict[str, Any]) -> None:
        """保存版本資訊"""
        try:
            with open(self.version_file, 'w', encoding='utf-8') as f:
                json.dump(version_info, f, indent=2, ensure_ascii=False)
            logger.info(f"版本資訊已保存: {self.version_file}")
        except Exception as e:
            logger.error(f"保存版本資訊失敗: {e}")
    
    def get_version_string(self) -> str:
        """獲取版本字串"""
        info = self.version_info
        return f"v{info['major']}.{info['minor']}.{info['patch']}-{info['build']}"
    
    def increment_version(self, version_type: str = "patch", description: str = "") -> str:
        """增加版本號碼"""
        if version_type == "major":
            self.version_info["major"] += 1
            self.version_info["minor"] = 0
            self.version_info["patch"] = 0
        elif version_type == "minor":
            self.version_info["minor"] += 1
            self.version_info["patch"] = 0
        elif version_type == "patch":
            self.version_info["patch"] += 1
        
        self.version_info["build"] += 1
        self.version_info["release_date"] = datetime.now().isoformat()
        
        if description:
            self.version_info["description"] = description
        
        self.version_info["changes"].append(f"{self.get_version_string()}: {description}")
        
        self.save_version_info(self.version_info)
        logger.info(f"版本已更新: {self.get_version_string()}")
        return self.get_version_string()
    
    def get_version_details(self) -> Dict[str, Any]:
        """獲取詳細版本資訊"""
        return {
            "version_string": self.get_version_string(),
            "major": self.version_info["major"],
            "minor": self.version_info["minor"],
            "patch": self.version_info["patch"],
            "build": self.version_info["build"],
            "release_date": self.version_info["release_date"],
            "description": self.version_info["description"],
            "changes": self.version_info["changes"]
        }


class VoltageMeasurement:
    """電壓量測類"""
    
    def __init__(self):
        self.measurement_config = {
            "voltage_range": (0.0, 5.0),  # 電壓範圍 (V)
            "accuracy": 0.01,  # 精度 (V)
            "sampling_rate": 1000,  # 採樣率 (Hz)
            "measurement_time": 1.0  # 量測時間 (秒)
        }
        self.measurement_history = []
    
    def measure_voltage(self, channel: int = 1, duration: float = None) -> Dict[str, Any]:
        """量測電壓"""
        if duration is None:
            duration = self.measurement_config["measurement_time"]
        
        try:
            # 模擬電壓量測（實際應用中這裡會連接真實的硬體）
            start_time = time.time()
            measurements = []
            
            # 模擬採樣數據
            num_samples = int(duration * self.measurement_config["sampling_rate"])
            for i in range(num_samples):
                # 模擬真實電壓值（帶有噪聲）
                base_voltage = random.uniform(3.0, 3.6)  # 模擬3.3V電源
                noise = random.gauss(0, 0.005)  # 高斯噪聲
                voltage = base_voltage + noise
                
                timestamp = start_time + (i / self.measurement_config["sampling_rate"])
                measurements.append({
                    "timestamp": timestamp,
                    "voltage": round(voltage, 4),
                    "channel": channel
                })
            
            end_time = time.time()
            
            # 計算統計數據
            voltages = [m["voltage"] for m in measurements]
            avg_voltage = sum(voltages) / len(voltages)
            min_voltage = min(voltages)
            max_voltage = max(voltages)
            std_voltage = (sum((v - avg_voltage) ** 2 for v in voltages) / len(voltages)) ** 0.5
            
            result = {
                "channel": channel,
                "duration": duration,
                "sampling_rate": self.measurement_config["sampling_rate"],
                "num_samples": num_samples,
                "measurements": measurements,
                "statistics": {
                    "average_voltage": round(avg_voltage, 4),
                    "min_voltage": round(min_voltage, 4),
                    "max_voltage": round(max_voltage, 4),
                    "std_voltage": round(std_voltage, 4),
                    "voltage_range": round(max_voltage - min_voltage, 4)
                },
                "timestamp": datetime.now().isoformat(),
                "status": "success"
            }
            
            self.measurement_history.append(result)
            logger.info(f"電壓量測完成 - 通道{channel}: 平均電壓={avg_voltage:.4f}V")
            
            return result
            
        except Exception as e:
            logger.error(f"電壓量測失敗 - 通道{channel}: {e}")
            return {
                "channel": channel,
                "duration": duration,
                "timestamp": datetime.now().isoformat(),
                "status": "error",
                "error": str(e)
            }
    
    def get_measurement_history(self) -> List[Dict[str, Any]]:
        """獲取量測歷史"""
        return self.measurement_history
    
    def clear_history(self) -> None:
        """清除量測歷史"""
        self.measurement_history.clear()
        logger.info("電壓量測歷史已清除")


class BluetoothMeasurement:
    """藍牙量測類"""
    
    def __init__(self):
        self.measurement_config = {
            "scan_duration": 10.0,  # 掃描時間 (秒)
            "rssi_threshold": -80,  # RSSI閾值 (dBm)
            "connection_timeout": 30.0,  # 連接超時 (秒)
            "supported_profiles": ["HFP", "A2DP", "AVRCP", "GATT"]
        }
        self.measurement_history = []
    
    def scan_bluetooth_devices(self, duration: float = None) -> Dict[str, Any]:
        """掃描藍牙設備"""
        if duration is None:
            duration = self.measurement_config["scan_duration"]
        
        try:
            # 模擬藍牙掃描（實際應用中這裡會使用真實的藍牙API）
            start_time = time.time()
            devices = []
            
            # 模擬發現的藍牙設備
            device_names = [
                "iPhone 15", "Samsung Galaxy", "MacBook Pro", "AirPods Pro",
                "Bose QC35", "JBL Flip", "Sony WH-1000XM4", "Beats Studio"
            ]
            
            num_devices = random.randint(3, 8)
            for i in range(num_devices):
                device = {
                    "name": random.choice(device_names),
                    "address": f"{random.randint(0, 255):02X}:{random.randint(0, 255):02X}:{random.randint(0, 255):02X}:{random.randint(0, 255):02X}:{random.randint(0, 255):02X}:{random.randint(0, 255):02X}",
                    "rssi": random.randint(-90, -30),
                    "device_class": random.choice(["Phone", "Computer", "Audio", "Wearable", "Other"]),
                    "supported_profiles": random.sample(self.measurement_config["supported_profiles"], 
                                                     random.randint(1, 3)),
                    "discovery_time": start_time + random.uniform(0, duration)
                }
                devices.append(device)
            
            end_time = time.time()
            
            # 計算統計數據
            rssi_values = [d["rssi"] for d in devices]
            avg_rssi = sum(rssi_values) / len(rssi_values)
            strong_signals = len([r for r in rssi_values if r > self.measurement_config["rssi_threshold"]])
            
            result = {
                "scan_duration": duration,
                "devices_found": len(devices),
                "devices": devices,
                "statistics": {
                    "average_rssi": round(avg_rssi, 2),
                    "strong_signals": strong_signals,
                    "weak_signals": len(devices) - strong_signals,
                    "device_types": list(set(d["device_class"] for d in devices))
                },
                "timestamp": datetime.now().isoformat(),
                "status": "success"
            }
            
            self.measurement_history.append(result)
            logger.info(f"藍牙掃描完成: 發現{len(devices)}個設備")
            
            return result
            
        except Exception as e:
            logger.error(f"藍牙掃描失敗: {e}")
            return {
                "scan_duration": duration,
                "timestamp": datetime.now().isoformat(),
                "status": "error",
                "error": str(e)
            }
    
    def test_bluetooth_connection(self, device_address: str, device_name: str) -> Dict[str, Any]:
        """測試藍牙連接"""
        try:
            # 模擬藍牙連接測試
            connection_time = random.uniform(1.0, 5.0)
            connection_success = random.random() > 0.1  # 90%成功率
            
            if connection_success:
                # 模擬連接成功後的測試
                profile_tests = {}
                for profile in self.measurement_config["supported_profiles"]:
                    profile_tests[profile] = {
                        "status": "success" if random.random() > 0.2 else "failed",
                        "response_time": random.uniform(0.1, 2.0),
                        "data_rate": random.uniform(1.0, 10.0)
                    }
                
                result = {
                    "device_address": device_address,
                    "device_name": device_name,
                    "connection_status": "success",
                    "connection_time": round(connection_time, 2),
                    "profile_tests": profile_tests,
                    "timestamp": datetime.now().isoformat(),
                    "status": "success"
                }
            else:
                result = {
                    "device_address": device_address,
                    "device_name": device_name,
                    "connection_status": "failed",
                    "connection_time": connection_time,
                    "error": "Connection timeout",
                    "timestamp": datetime.now().isoformat(),
                    "status": "error"
                }
            
            self.measurement_history.append(result)
            logger.info(f"藍牙連接測試完成: {device_name} - {result['connection_status']}")
            
            return result
            
        except Exception as e:
            logger.error(f"藍牙連接測試失敗: {e}")
            return {
                "device_address": device_address,
                "device_name": device_name,
                "timestamp": datetime.now().isoformat(),
                "status": "error",
                "error": str(e)
            }
    
    def get_measurement_history(self) -> List[Dict[str, Any]]:
        """獲取量測歷史"""
        return self.measurement_history
    
    def clear_history(self) -> None:
        """清除量測歷史"""
        self.measurement_history.clear()
        logger.info("藍牙量測歷史已清除")


class TestReportGenerator:
    """測試報告生成器"""
    
    def __init__(self, output_dir: str = "test_reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def generate_csv_report(self, voltage_data: List[Dict[str, Any]], 
                           bluetooth_data: List[Dict[str, Any]], 
                           version_info: Dict[str, Any]) -> str:
        """生成CSV格式的測試報告"""
        try:
            # 準備電壓數據
            voltage_rows = []
            for measurement in voltage_data:
                if measurement.get("status") == "success":
                    stats = measurement["statistics"]
                    voltage_rows.append({
                        "測試類型": "電壓量測",
                        "通道": measurement["channel"],
                        "平均電壓(V)": stats["average_voltage"],
                        "最小電壓(V)": stats["min_voltage"],
                        "最大電壓(V)": stats["max_voltage"],
                        "標準差(V)": stats["std_voltage"],
                        "電壓範圍(V)": stats["voltage_range"],
                        "採樣數量": measurement["num_samples"],
                        "量測時間(秒)": measurement["duration"],
                        "狀態": measurement["status"],
                        "時間戳": measurement["timestamp"]
                    })
            
            # 準備藍牙數據
            bluetooth_rows = []
            for measurement in bluetooth_data:
                if measurement.get("status") == "success":
                    if "devices" in measurement:  # 掃描結果
                        for device in measurement["devices"]:
                            bluetooth_rows.append({
                                "測試類型": "藍牙掃描",
                                "設備名稱": device["name"],
                                "設備地址": device["address"],
                                "RSSI(dBm)": device["rssi"],
                                "設備類型": device["device_class"],
                                "支援協議": ", ".join(device["supported_profiles"]),
                                "發現時間": device["discovery_time"],
                                "掃描時間(秒)": measurement["scan_duration"],
                                "狀態": measurement["status"],
                                "時間戳": measurement["timestamp"]
                            })
                    elif "connection_status" in measurement:  # 連接測試
                        bluetooth_rows.append({
                            "測試類型": "藍牙連接",
                            "設備名稱": measurement["device_name"],
                            "設備地址": measurement["device_address"],
                            "連接狀態": measurement["connection_status"],
                            "連接時間(秒)": measurement.get("connection_time", 0),
                            "協議測試": json.dumps(measurement.get("profile_tests", {}), ensure_ascii=False),
                            "狀態": measurement["status"],
                            "時間戳": measurement["timestamp"]
                        })
            
            # 合併所有數據
            all_rows = voltage_rows + bluetooth_rows
            
            if all_rows:
                # 創建DataFrame
                df = pd.DataFrame(all_rows)
                
                # 生成檔案名
                filename = f"test_report_{self.timestamp}.csv"
                filepath = self.output_dir / filename
                
                # 保存CSV
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
                logger.info(f"CSV報告已生成: {filepath}")
                
                return str(filepath)
            else:
                logger.warning("沒有有效的測試數據可生成報告")
                return ""
                
        except Exception as e:
            logger.error(f"生成CSV報告失敗: {e}")
            return ""
    
    def generate_excel_report(self, voltage_data: List[Dict[str, Any]], 
                             bluetooth_data: List[Dict[str, Any]], 
                             version_info: Dict[str, Any]) -> str:
        """生成Excel格式的測試報告"""
        try:
            # 創建Excel工作簿
            wb = openpyxl.Workbook()
            
            # 移除預設工作表
            wb.remove(wb.active)
            
            # 創建版本資訊工作表
            self._create_version_sheet(wb, version_info)
            
            # 創建電壓量測工作表
            self._create_voltage_sheet(wb, voltage_data)
            
            # 創建藍牙量測工作表
            self._create_bluetooth_sheet(wb, bluetooth_data)
            
            # 創建摘要工作表
            self._create_summary_sheet(wb, voltage_data, bluetooth_data, version_info)
            
            # 生成檔案名
            filename = f"test_report_{self.timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            # 保存Excel
            wb.save(filepath)
            logger.info(f"Excel報告已生成: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"生成Excel報告失敗: {e}")
            return ""
    
    def _create_version_sheet(self, wb: openpyxl.Workbook, version_info: Dict[str, Any]) -> None:
        """創建版本資訊工作表"""
        ws = wb.create_sheet("版本資訊")
        
        # 標題樣式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 設置標題
        ws['A1'] = "測試腳本版本資訊"
        ws['A1'].font = title_font
        ws.merge_cells('A1:C1')
        
        # 版本詳細資訊
        headers = ["項目", "值", "說明"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 版本數據
        version_data = [
            ["版本號碼", version_info["version_string"], "當前腳本版本"],
            ["主版本號", version_info["major"], "主要功能更新"],
            ["次版本號", version_info["minor"], "功能增強"],
            ["修補版本", version_info["patch"], "錯誤修復"],
            ["建置編號", version_info["build"], "建置序號"],
            ["發布日期", version_info["release_date"], "版本發布時間"],
            ["版本描述", version_info["description"], "版本主要特點"]
        ]
        
        for row, (item, value, desc) in enumerate(version_data, 4):
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=desc)
        
        # 變更歷史
        ws['A12'] = "變更歷史"
        ws['A12'].font = title_font
        
        for row, change in enumerate(version_info["changes"], 13):
            ws.cell(row=row, column=1, value=change)
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
    
    def _create_voltage_sheet(self, wb: openpyxl.Workbook, voltage_data: List[Dict[str, Any]]) -> None:
        """創建電壓量測工作表"""
        ws = wb.create_sheet("電壓量測")
        
        # 標題樣式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # 設置標題
        ws['A1'] = "電壓量測結果"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        
        # 表頭
        headers = [
            "通道", "平均電壓(V)", "最小電壓(V)", "最大電壓(V)", 
            "標準差(V)", "電壓範圍(V)", "採樣數量", "量測時間(秒)", "狀態", "時間戳"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 數據行
        row = 4
        for measurement in voltage_data:
            if measurement.get("status") == "success":
                stats = measurement["statistics"]
                data = [
                    measurement["channel"],
                    stats["average_voltage"],
                    stats["min_voltage"],
                    stats["max_voltage"],
                    stats["std_voltage"],
                    stats["voltage_range"],
                    measurement["num_samples"],
                    measurement["duration"],
                    measurement["status"],
                    measurement["timestamp"]
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
        
        # 調整欄寬
        for col in range(1, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    def _create_bluetooth_sheet(self, wb: openpyxl.Workbook, bluetooth_data: List[Dict[str, Any]]) -> None:
        """創建藍牙量測工作表"""
        ws = wb.create_sheet("藍牙量測")
        
        # 標題樣式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        
        # 設置標題
        ws['A1'] = "藍牙量測結果"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        
        # 表頭
        headers = [
            "測試類型", "設備名稱", "設備地址", "RSSI(dBm)", 
            "設備類型", "支援協議", "連接狀態", "時間戳"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 數據行
        row = 4
        for measurement in bluetooth_data:
            if measurement.get("status") == "success":
                if "devices" in measurement:  # 掃描結果
                    for device in measurement["devices"]:
                        data = [
                            "藍牙掃描",
                            device["name"],
                            device["address"],
                            device["rssi"],
                            device["device_class"],
                            ", ".join(device["supported_profiles"]),
                            "N/A",
                            measurement["timestamp"]
                        ]
                        
                        for col, value in enumerate(data, 1):
                            ws.cell(row=row, column=col, value=value)
                        
                        row += 1
                
                elif "connection_status" in measurement:  # 連接測試
                    data = [
                        "藍牙連接",
                        measurement["device_name"],
                        measurement["device_address"],
                        "N/A",
                        "N/A",
                        "N/A",
                        measurement["connection_status"],
                        measurement["timestamp"]
                    ]
                    
                    for col, value in enumerate(data, 1):
                        ws.cell(row=row, column=col, value=value)
                    
                    row += 1
        
        # 調整欄寬
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, voltage_data: List[Dict[str, Any]], 
                             bluetooth_data: List[Dict[str, Any]], version_info: Dict[str, Any]) -> None:
        """創建摘要工作表"""
        ws = wb.create_sheet("測試摘要")
        
        # 標題樣式
        title_font = Font(bold=True, size=16)
        subtitle_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # 設置標題
        ws['A1'] = "測試報告摘要"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # 測試概況
        ws['A3'] = "測試概況"
        ws['A3'].font = subtitle_font
        
        # 統計數據
        voltage_count = len([d for d in voltage_data if d.get("status") == "success"])
        bluetooth_count = len([d for d in bluetooth_data if d.get("status") == "success"])
        
        summary_data = [
            ["測試項目", "數量", "狀態", "說明"],
            ["電壓量測", voltage_count, "完成" if voltage_count > 0 else "無數據", "電壓通道量測"],
            ["藍牙掃描", bluetooth_count, "完成" if bluetooth_count > 0 else "無數據", "藍牙設備掃描"],
            ["藍牙連接", bluetooth_count, "完成" if bluetooth_count > 0 else "無數據", "藍牙連接測試"],
            ["腳本版本", version_info["version_string"], "當前版本", "測試腳本版本號碼"]
        ]
        
        for row, (item, count, status, desc) in enumerate(summary_data, 4):
            for col, value in enumerate([item, count, status, desc], 1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 4:  # 表頭
                    cell.font = header_font
                    cell.fill = header_fill
        
        # 測試時間
        ws['A10'] = "測試時間"
        ws['A10'].font = subtitle_font
        ws['B10'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30


class TestMeasurementScript:
    """測試量測腳本主類"""
    
    def __init__(self):
        self.version_manager = VersionManager()
        self.voltage_measurement = VoltageMeasurement()
        self.bluetooth_measurement = BluetoothMeasurement()
        self.report_generator = TestReportGenerator()
        
        logger.info(f"測試腳本已初始化，版本: {self.version_manager.get_version_string()}")
    
    def run_voltage_tests(self, channels: List[int] = None, duration: float = 1.0) -> List[Dict[str, Any]]:
        """執行電壓量測測試"""
        if channels is None:
            channels = [1, 2, 3]  # 預設測試3個通道
        
        logger.info(f"開始電壓量測測試，通道: {channels}, 持續時間: {duration}秒")
        
        results = []
        for channel in channels:
            logger.info(f"測試通道 {channel}")
            result = self.voltage_measurement.measure_voltage(channel, duration)
            results.append(result)
            
            # 等待一下再測試下一個通道
            time.sleep(0.5)
        
        logger.info(f"電壓量測測試完成，共測試 {len(channels)} 個通道")
        return results
    
    def run_bluetooth_tests(self, scan_duration: float = 10.0, test_connections: bool = True) -> List[Dict[str, Any]]:
        """執行藍牙量測測試"""
        logger.info(f"開始藍牙量測測試，掃描時間: {scan_duration}秒")
        
        results = []
        
        # 藍牙掃描
        logger.info("執行藍牙設備掃描...")
        scan_result = self.bluetooth_measurement.scan_bluetooth_devices(scan_duration)
        results.append(scan_result)
        
        # 藍牙連接測試
        if test_connections and scan_result.get("status") == "success":
            devices = scan_result.get("devices", [])
            if devices:
                logger.info(f"開始測試藍牙連接，共 {len(devices)} 個設備")
                
                # 選擇前3個設備進行連接測試
                test_devices = devices[:3]
                for device in test_devices:
                    logger.info(f"測試連接設備: {device['name']}")
                    connection_result = self.bluetooth_measurement.test_bluetooth_connection(
                        device["address"], device["name"]
                    )
                    results.append(connection_result)
                    time.sleep(1)  # 等待連接完成
        
        logger.info("藍牙量測測試完成")
        return results
    
    def generate_reports(self, voltage_results: List[Dict[str, Any]], 
                        bluetooth_results: List[Dict[str, Any]], 
                        output_formats: List[str] = None) -> Dict[str, str]:
        """生成測試報告"""
        if output_formats is None:
            output_formats = ["csv", "xlsx"]
        
        version_info = self.version_manager.get_version_details()
        report_files = {}
        
        logger.info("開始生成測試報告...")
        
        if "csv" in output_formats:
            csv_file = self.report_generator.generate_csv_report(
                voltage_results, bluetooth_results, version_info
            )
            if csv_file:
                report_files["csv"] = csv_file
        
        if "xlsx" in output_formats:
            xlsx_file = self.report_generator.generate_excel_report(
                voltage_results, bluetooth_results, version_info
            )
            if xlsx_file:
                report_files["xlsx"] = xlsx_file
        
        logger.info(f"測試報告生成完成: {list(report_files.keys())}")
        return report_files
    
    def run_full_test_suite(self) -> Dict[str, Any]:
        """執行完整的測試套件"""
        logger.info("開始執行完整測試套件...")
        
        # 執行電壓測試
        voltage_results = self.run_voltage_tests()
        
        # 執行藍牙測試
        bluetooth_results = self.run_bluetooth_tests()
        
        # 生成報告
        report_files = self.generate_reports(voltage_results, bluetooth_results)
        
        # 準備測試摘要
        test_summary = {
            "test_timestamp": datetime.now().isoformat(),
            "script_version": self.version_manager.get_version_string(),
            "voltage_tests": {
                "total_channels": len(voltage_results),
                "successful_tests": len([r for r in voltage_results if r.get("status") == "success"]),
                "failed_tests": len([r for r in voltage_results if r.get("status") == "error"])
            },
            "bluetooth_tests": {
                "total_tests": len(bluetooth_results),
                "successful_tests": len([r for r in bluetooth_results if r.get("status") == "success"]),
                "failed_tests": len([r for r in bluetooth_results if r.get("status") == "error"])
            },
            "report_files": report_files
        }
        
        logger.info("完整測試套件執行完成")
        return test_summary


def main():
    """主函數"""
    try:
        print("=" * 60)
        print("測試量測腳本 v1.0.0")
        print("=" * 60)
        
        # 創建測試腳本實例
        test_script = TestMeasurementScript()
        
        # 顯示版本資訊
        version_info = test_script.version_manager.get_version_details()
        print(f"腳本版本: {version_info['version_string']}")
        print(f"發布日期: {version_info['release_date']}")
        print(f"版本描述: {version_info['description']}")
        print()
        
        # 執行完整測試套件
        print("開始執行測試...")
        test_summary = test_script.run_full_test_suite()
        
        # 顯示測試結果
        print("\n" + "=" * 60)
        print("測試結果摘要")
        print("=" * 60)
        print(f"測試時間: {test_summary['test_timestamp']}")
        print(f"腳本版本: {test_summary['script_version']}")
        print()
        
        print("電壓測試結果:")
        print(f"  總通道數: {test_summary['voltage_tests']['total_channels']}")
        print(f"  成功測試: {test_summary['voltage_tests']['successful_tests']}")
        print(f"  失敗測試: {test_summary['voltage_tests']['failed_tests']}")
        print()
        
        print("藍牙測試結果:")
        print(f"  總測試數: {test_summary['bluetooth_tests']['total_tests']}")
        print(f"  成功測試: {test_summary['bluetooth_tests']['successful_tests']}")
        print(f"  失敗測試: {test_summary['bluetooth_tests']['failed_tests']}")
        print()
        
        print("報告檔案:")
        for format_type, file_path in test_summary['report_files'].items():
            print(f"  {format_type.upper()}: {file_path}")
        
        print("\n測試完成！")
        
    except KeyboardInterrupt:
        print("\n\n測試被使用者中斷")
    except Exception as e:
        logger.error(f"主程式執行失敗: {e}")
        print(f"\n程式執行失敗: {e}")
        print("請檢查日誌檔案獲取詳細錯誤資訊")


if __name__ == "__main__":
    main() 