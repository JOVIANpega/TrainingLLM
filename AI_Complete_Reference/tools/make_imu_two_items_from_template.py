# -*- coding: utf-8 -*-
"""
make_imu_two_items_from_template.py

用途:
    讀取 `Centimani DOC/scripe_EXCEL/VALO360/VALO360_TestFlow_20250326_IMU.xlsx` 作為模板，
    在 `IMU` 分頁只保留兩個測項：
      1) Check E compass
      2) Check E compass result
    並輸出到 `test script/System/VALO360_IMU_TwoItems.xlsx`。

注意:
    - 僅修改 `IMU` 分頁的資料列，其他分頁保留不變
    - 依現有表頭自動對應欄位（Index 或 Index_*、Phase、Title、Send、Reply、Ref、Validation、Waiting、ExecMode、TestID 等）
"""
from pathlib import Path
from typing import Dict, List
from openpyxl import load_workbook


def build_header_map(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        header = str(v).strip()
        header_map[header] = c
    return header_map


def find_index_col(header_map: Dict[str, int]) -> int:
    # 支援 Index 或 Index_*
    for k, c in header_map.items():
        if k.lower().startswith('index'):
            return c
    return 0


def clear_rows_keep_header(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def set_cell_if_exists(ws, header_map: Dict[str, int], row_idx: int, key: str, value) -> None:
    if key in header_map:
        ws.cell(row_idx, header_map[key], value)


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    template = root / 'Centimani DOC' / 'scripe_EXCEL' / 'VALO360' / 'VALO360_TestFlow_20250326_IMU.xlsx'
    if not template.exists():
        raise SystemExit(f'找不到模板: {template}')

    wb = load_workbook(template)
    ws = wb['IMU'] if 'IMU' in wb.sheetnames else wb.active

    header_map = build_header_map(ws)
    index_col = find_index_col(header_map)

    # 僅保留表頭
    clear_rows_keep_header(ws)

    # 寫入兩個測項（Index=5,6；Phase=0）
    # Row 2: Index 5
    r = 2
    if index_col:
        ws.cell(r, index_col, 5)
    set_cell_if_exists(ws, header_map, r, 'Phase', 0)
    set_cell_if_exists(ws, header_map, r, 'Title', 'Check E compass')
    set_cell_if_exists(ws, header_map, r, 'Send', 'UART1@imu mag selftest')
    set_cell_if_exists(ws, header_map, r, 'Reply', '@SELFTEST:PASS')
    set_cell_if_exists(ws, header_map, r, 'Waiting', '3000')
    set_cell_if_exists(ws, header_map, r, 'ExecMode', '0')
    set_cell_if_exists(ws, header_map, r, 'Description', '啟動磁力計自檢，等待 PASS')
    set_cell_if_exists(ws, header_map, r, 'TestID', '<查表填入>')

    # Row 3: Index 6
    r = 3
    if index_col:
        ws.cell(r, index_col, 6)
    set_cell_if_exists(ws, header_map, r, 'Phase', 0)
    set_cell_if_exists(ws, header_map, r, 'Title', 'Check E compass result')
    set_cell_if_exists(ws, header_map, r, 'Send', 'UART1@imu mag status')
    set_cell_if_exists(ws, header_map, r, 'Reply', '@ECOMPASS:PASS')
    # 若有提供數值可加：Ref/Validation（此處留白）
    set_cell_if_exists(ws, header_map, r, 'Waiting', '3000')
    set_cell_if_exists(ws, header_map, r, 'ExecMode', '0')
    set_cell_if_exists(ws, header_map, r, 'Description', '讀取磁羅盤狀態，驗證 PASS')
    set_cell_if_exists(ws, header_map, r, 'TestID', '<查表填入>')

    # 輸出
    out_dir = root / 'test script' / 'System'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / 'VALO360_IMU_TwoItems.xlsx'
    wb.save(out_path)
    print(f'已輸出: {out_path}')


if __name__ == '__main__':
    main() 