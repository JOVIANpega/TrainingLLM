#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析Excel範例檔案結構

用於分析Centimani DOC中的Excel範例檔案，了解正確的格式要求
"""

import os
import pandas as pd
import openpyxl
from pathlib import Path
import json

def analyze_excel_file(file_path):
    """分析Excel檔案結構"""
    try:
        # 讀取Excel檔案
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        analysis = {
            "file_name": os.path.basename(file_path),
            "file_size": os.path.getsize(file_path),
            "sheets": [],
            "total_sheets": len(wb.sheetnames)
        }
        
        # 分析每個工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 獲取工作表資訊
            sheet_info = {
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimensions": f"{ws.max_row} x {ws.max_column}",
                "sample_data": []
            }
            
            # 讀取前幾行數據作為範例
            for row in range(1, min(6, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(11, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                sheet_info["sample_data"].append(row_data)
            
            analysis["sheets"].append(sheet_info)
        
        wb.close()
        return analysis
        
    except Exception as e:
        return {
            "file_name": os.path.basename(file_path),
            "error": str(e)
        }

def analyze_directory(directory_path):
    """分析目錄中的所有Excel檔案"""
    excel_files = []
    
    for file_path in Path(directory_path).rglob("*.xlsx"):
        if file_path.is_file():
            print(f"分析檔案: {file_path}")
            analysis = analyze_excel_file(str(file_path))
            excel_files.append(analysis)
    
    return excel_files

def main():
    """主函數"""
    # 分析KANGAROO目錄
    kangaroo_dir = "../../Centimani DOC/scripe_EXCEL/KANGAROO"
    print("=" * 60)
    print("分析KANGAROO Excel範例檔案")
    print("=" * 60)
    
    kangaroo_files = analyze_directory(kangaroo_dir)
    
    # 分析VALO360目錄
    valo360_dir = "../../Centimani DOC/scripe_EXCEL/VALO360"
    print("\n" + "=" * 60)
    print("分析VALO360 Excel範例檔案")
    print("=" * 60)
    
    valo360_files = analyze_directory(valo360_dir)
    
    # 保存分析結果
    all_analysis = {
        "kangaroo_files": kangaroo_files,
        "valo360_files": valo360_files,
        "analysis_time": pd.Timestamp.now().isoformat()
    }
    
    # 保存為JSON檔案
    with open("excel_templates_analysis.json", "w", encoding="utf-8") as f:
        json.dump(all_analysis, f, indent=2, ensure_ascii=False)
    
    print(f"\n分析完成！結果已保存到: excel_templates_analysis.json")
    
    # 顯示摘要
    print(f"\nKANGAROO檔案數量: {len(kangaroo_files)}")
    print(f"VALO360檔案數量: {len(valo360_files)}")
    
    # 顯示幾個重要檔案的結構
    print("\n重要檔案結構摘要:")
    for category, files in [("KANGAROO", kangaroo_files), ("VALO360", valo360_files)]:
        print(f"\n{category}:")
        for file_info in files[:3]:  # 只顯示前3個
            if "sheets" in file_info:
                print(f"  {file_info['file_name']}: {len(file_info['sheets'])} 工作表")
                for sheet in file_info['sheets']:
                    print(f"    - {sheet['name']}: {sheet['dimensions']}")

if __name__ == "__main__":
    main() 