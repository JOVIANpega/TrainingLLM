#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
驗證Centimani格式測試腳本

用於驗證生成的Excel檔案是否符合Centimani格式要求
"""

import openpyxl
import os
from pathlib import Path

def verify_excel_file(file_path):
    """驗證Excel檔案結構"""
    try:
        wb = openpyxl.load_workbook(file_path)
        
        print(f"驗證檔案: {os.path.basename(file_path)}")
        print("=" * 60)
        
        # 檢查工作表
        expected_sheets = ["Properties", "Switch", "Instrument", "DUTs", "Test_Items"]
        actual_sheets = wb.sheetnames
        
        print(f"工作表檢查:")
        for sheet in expected_sheets:
            if sheet in actual_sheets:
                print(f"  ✅ {sheet}")
            else:
                print(f"  ❌ {sheet} (缺失)")
        
        print()
        
        # 檢查每個工作表的內容
        for sheet_name in expected_sheets:
            if sheet_name in actual_sheets:
                ws = wb[sheet_name]
                print(f"工作表: {sheet_name}")
                print(f"  尺寸: {ws.max_row} x {ws.max_column}")
                
                # 檢查標題行
                if ws.max_row > 0:
                    headers = []
                    for col in range(1, min(ws.max_column + 1, 11)):  # 最多檢查10列
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value:
                            headers.append(str(cell_value))
                    
                    print(f"  標題: {', '.join(headers[:5])}{'...' if len(headers) > 5 else ''}")
                    
                    # 檢查數據行
                    if ws.max_row > 1:
                        data_rows = ws.max_row - 1
                        print(f"  數據行數: {data_rows}")
                        
                        # 顯示前幾行數據
                        for row in range(2, min(6, ws.max_row + 1)):
                            row_data = []
                            for col in range(1, min(ws.max_column + 1, 6)):  # 最多顯示5列
                                cell_value = ws.cell(row=row, column=col).value
                                row_data.append(str(cell_value) if cell_value is not None else "")
                            print(f"    行{row}: {', '.join(row_data)}")
                
                print()
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"驗證失敗: {e}")
        return False

def main():
    """主函數"""
    print("Centimani格式測試腳本驗證器")
    print("=" * 60)
    
    # 查找生成的Centimani腳本檔案
    script_files = []
    for file_path in Path(".").glob("Centimani_TestScript_*.xlsx"):
        script_files.append(str(file_path))
    
    if not script_files:
        print("未找到Centimani測試腳本檔案")
        return
    
    print(f"找到 {len(script_files)} 個腳本檔案:")
    for file_path in script_files:
        print(f"  {file_path}")
    
    print()
    
    # 驗證每個檔案
    for file_path in script_files:
        verify_excel_file(file_path)
        print("-" * 60)
        print()

if __name__ == "__main__":
    main() 