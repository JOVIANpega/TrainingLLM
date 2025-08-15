#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
腳本內容修改工具
用於實際修改Excel腳本內容，刪除不需要的測試項目，只保留目標功能
"""

import openpyxl
import os
import sys
from pathlib import Path

def modify_login_script(input_file, output_file):
    """修改LOGIN測試腳本，只保留登入相關測試"""
    print(f"正在修改LOGIN腳本: {input_file}")
    
    # 載入工作簿
    wb = openpyxl.load_workbook(input_file)
    
    # 修改Properties工作表
    if 'Properties' in wb.sheetnames:
        ws = wb['Properties']
        # 更新專案資訊
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and 'VALO360' in str(cell.value):
                    cell.value = str(cell.value).replace('VALO360', 'LOGIN_TEST')
                if cell.value and '主機板功能測試' in str(cell.value):
                    cell.value = 'LOGIN登入功能測試'
    
    # 修改MB工作表
    if 'MB' in wb.sheetnames:
        ws = wb['MB']
        rows_to_delete = []
        
        # 找到需要保留的行（Phase 0, 1, 97）
        keep_rows = []
        for row_num in range(2, ws.max_row + 1):
            phase_cell = ws.cell(row=row_num, column=2)  # Phase欄位
            if phase_cell.value in ['0', '1', '97']:
                keep_rows.append(row_num)
            else:
                rows_to_delete.append(row_num)
        
        # 從後往前刪除行，避免索引變化
        for row_num in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_num)
        
        # 修改Phase 1的內容為LOGIN測試
        for row_num in keep_rows:
            phase_cell = ws.cell(row=row_num, column=2)
            if phase_cell.value == '1':
                # 修改Title
                title_cell = ws.cell(row=row_num, column=3)
                title_cell.value = 'Test Login'
                
                # 修改Send
                send_cell = ws.cell(row=row_num, column=7)
                send_cell.value = 'SSH1@login'
                
                # 修改Reply
                reply_cell = ws.cell(row=row_num, column=8)
                reply_cell.value = '@Login:'
                
                # 修改TestID
                testid_cell = ws.cell(row=row_num, column=16)
                testid_cell.value = 'L001'
                
                # 修改Description
                desc_cell = ws.cell(row=row_num, column=15)
                desc_cell.value = 'Test login functionality'
        
        print(f"已刪除 {len(rows_to_delete)} 行，保留 {len(keep_rows)} 行")
    
    # 保存修改後的腳本
    wb.save(output_file)
    print(f"LOGIN腳本修改完成: {output_file}")

def modify_voltage_script(input_file, output_file):
    """修改電壓測試腳本，只保留電壓相關測試"""
    print(f"正在修改電壓腳本: {input_file}")
    
    # 載入工作簿
    wb = openpyxl.load_workbook(input_file)
    
    # 修改Properties工作表
    if 'Properties' in wb.sheetnames:
        ws = wb['Properties']
        # 更新專案資訊
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and 'VALO360' in str(cell.value):
                    cell.value = str(cell.value).replace('VALO360', 'VOLTAGE_TEST')
                if cell.value and '主機板功能測試' in str(cell.value):
                    cell.value = '電壓測試腳本'
    
    # 修改MB工作表
    if 'MB' in wb.sheetnames:
        ws = wb['MB']
        rows_to_delete = []
        
        # 找到需要保留的行（Phase 0, 1, 97）
        keep_rows = []
        for row_num in range(2, ws.max_row + 1):
            phase_cell = ws.cell(row=row_num, column=2)  # Phase欄位
            if phase_cell.value in ['0', '1', '97']:
                keep_rows.append(row_num)
            else:
                rows_to_delete.append(row_num)
        
        # 從後往前刪除行，避免索引變化
        for row_num in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_num)
        
        # 修改Phase 1的內容為電壓測試
        for row_num in keep_rows:
            phase_cell = ws.cell(row=row_num, column=2)
            if phase_cell.value == '1':
                # 修改Title
                title_cell = ws.cell(row=row_num, column=3)
                title_cell.value = 'Measure Voltage'
                
                # 修改Instrument
                instrument_cell = ws.cell(row=row_num, column=5)
                instrument_cell.value = 'DMM.MeasureV'
                
                # 修改Send
                send_cell = ws.cell(row=row_num, column=7)
                send_cell.value = ':DMM.MeasureV,CH1,DC'
                
                # 修改Reply
                reply_cell = ws.cell(row=row_num, column=8)
                reply_cell.value = '@Voltage:'
                
                # 修改Ref
                ref_cell = ws.cell(row=row_num, column=10)
                ref_cell.value = 'voltage'
                
                # 修改Validation
                validation_cell = ws.cell(row=row_num, column=11)
                validation_cell.value = '(voltage,%f,4.75,5.25)'
                
                # 修改TestID
                testid_cell = ws.cell(row=row_num, column=16)
                testid_cell.value = 'V001'
                
                # 修改Description
                desc_cell = ws.cell(row=row_num, column=15)
                desc_cell.value = 'Measure 5V rail voltage'
        
        print(f"已刪除 {len(rows_to_delete)} 行，保留 {len(keep_rows)} 行")
    
    # 保存修改後的腳本
    wb.save(output_file)
    print(f"電壓腳本修改完成: {output_file}")

def main():
    """主函數"""
    print("腳本內容修改工具")
    print("=" * 50)
    
    # 檢查輸入檔案
    login_input = "LOGIN_Test_Correct.xlsx"
    voltage_input = "Voltage_Test_Correct.xlsx"
    
    if not os.path.exists(login_input):
        print(f"錯誤: 找不到輸入檔案 {login_input}")
        return
    
    if not os.path.exists(voltage_input):
        print(f"錯誤: 找不到輸入檔案 {voltage_input}")
        return
    
    # 修改LOGIN腳本
    login_output = "LOGIN_Test_Final.xlsx"
    modify_login_script(login_input, login_output)
    
    # 修改電壓腳本
    voltage_output = "Voltage_Test_Final.xlsx"
    modify_voltage_script(voltage_input, voltage_output)
    
    print("\n腳本修改完成！")
    print(f"LOGIN腳本: {login_output}")
    print(f"電壓腳本: {voltage_output}")

if __name__ == "__main__":
    main() 