#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
統一的Centimani腳本生成器
Unified Centimani Script Generator

整合了三個現有工具的所有功能：
1. ai_script_development_guide.py - AI腳本開發指南
2. generate_test_script.py - 測試腳本生成器
3. modify_script_from_template.py - 從模板修改腳本

主要功能：
1. 生成完整的測試腳本（包含所有工作表）
2. 從現有腳本模板修改生成新腳本
3. 生成AI腳本開發指南和最佳實踐
4. 支援多種腳本類型和專案
5. 自動生成INI配置檔案
6. 提供腳本驗證和檢查功能

作者: AI Assistant
版本: 2.0.0
創建日期: 2025-01-XX
更新日期: 2025-01-XX
"""

import os
import json
import sys
import shutil
import datetime
import configparser
import argparse
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from collections import defaultdict
import logging
from openpyxl import Workbook, load_workbook

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('unified_script_generator.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class UnifiedCentimaniScriptGenerator:
    """統一的Centimani腳本生成器"""
    
    def __init__(self, base_path: str = None):
        """
        初始化腳本生成器
        
        Args:
            base_path: 基礎路徑，預設為當前目錄
        """
        self.base_path = Path(base_path) if base_path else Path.cwd()
        self.timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 建立輸出目錄
        self.output_dir = self.base_path / "unified_script_output"
        self.output_dir.mkdir(exist_ok=True)
        
        # 專案分類
        self.project_categories = {
            "KANGAROO": "Kangaroo專案",
            "VALO360": "VALO360專案", 
            "RAPTOR": "Raptor專案",
            "Kilimanjaro": "Kilimanjaro專案",
            "ZEBRA": "Zebra專案"
        }
        
        # 腳本模板配置
        self.script_templates = {
            "basic": {
                "description": "基本三測項腳本（版本、電壓、BT）",
                "test_items": 3,
                "sheets": ["Properties", "Instrument", "Switch", "DUTs", "Main"]
            },
            "standard": {
                "description": "標準測試腳本（包含完整配置）",
                "test_items": 5,
                "sheets": ["Properties", "Instrument", "Switch", "DUTs", "Main", "Config"]
            },
            "advanced": {
                "description": "進階測試腳本（多站別、複雜邏輯）",
                "test_items": 10,
                "sheets": ["Properties", "Instrument", "Switch", "DUTs", "Main", "Config", "Macro"]
            }
        }
        
        logger.info(f"初始化統一的Centimani腳本生成器")
        logger.info(f"基礎路徑: {self.base_path}")
        logger.info(f"輸出目錄: {self.output_dir}")
    
    def generate_new_script(self, 
                           template_type: str = "basic",
                           station_name: str = "Main",
                           project_name: str = "TestScript",
                           output_path: str = None) -> str:
        """
        生成新的測試腳本
        
        Args:
            template_type: 模板類型 (basic, standard, advanced)
            station_name: 站別名稱
            project_name: 專案名稱
            output_path: 輸出路徑
            
        Returns:
            生成的腳本路徑
        """
        try:
            logger.info(f"開始生成新腳本: {template_type} 模板")
            
            if template_type not in self.script_templates:
                raise ValueError(f"不支援的模板類型: {template_type}")
            
            template_config = self.script_templates[template_type]
            
            # 建立輸出目錄結構
            if not output_path:
                output_path = self.output_dir / f"{project_name}_{template_type}_{self.timestamp}.xlsx"
            
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 建立工作簿
            wb = Workbook()
            
            # 生成各個工作表
            self._create_properties_sheet(wb, project_name, station_name)
            self._create_instrument_sheet(wb, station_name)
            self._create_switch_sheet(wb, station_name)
            self._create_duts_sheet(wb, station_name)
            self._create_main_test_sheet(wb, station_name, template_config)
            
            # 如果是進階模板，添加額外工作表
            if template_type == "advanced":
                self._create_config_sheet(wb, station_name)
                self._create_macro_sheet(wb, station_name)
            
            # 儲存腳本
            wb.save(str(output_path))
            logger.info(f"腳本生成完成: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"生成新腳本時發生錯誤: {e}")
            raise
    
    def modify_from_template(self, 
                           template_path: str,
                           output_path: str,
                           station_name: str = "Main",
                           modifications: Dict[str, Any] = None) -> str:
        """
        從現有模板修改生成腳本
        
        Args:
            template_path: 模板腳本路徑
            output_path: 輸出腳本路徑
            station_name: 站別名稱
            modifications: 修改配置
            
        Returns:
            修改後的腳本路徑
        """
        try:
            logger.info(f"開始從模板修改腳本: {template_path}")
            
            template_path = Path(template_path)
            output_path = Path(output_path)
            
            if not template_path.exists():
                raise FileNotFoundError(f"模板腳本不存在: {template_path}")
            
            # 載入模板工作簿
            wb = load_workbook(template_path)
            
            # 找到目標工作表
            target_sheet = self._find_target_sheet(wb, station_name)
            if not target_sheet:
                raise ValueError(f"找不到可用的測試腳本工作表: {station_name}")
            
            ws = wb[target_sheet]
            
            # 建立表頭對應
            header_map = self._build_header_map(ws)
            
            # 清除現有數據，保留表頭
            self._clear_rows_keep_header(ws)
            
            # 應用修改配置
            if modifications:
                self._apply_modifications(ws, header_map, modifications)
            else:
                # 使用預設的三測項配置
                self._apply_default_test_items(ws, header_map)
            
            # 儲存修改後的腳本
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output_path))
            
            logger.info(f"模板修改完成: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"從模板修改腳本時發生錯誤: {e}")
            raise
    
    def generate_development_guide(self, 
                                 project_data: Dict[str, Any] = None,
                                 output_path: str = None) -> str:
        """
        生成AI腳本開發指南
        
        Args:
            project_data: 專案資料
            output_path: 輸出路徑
            
        Returns:
            生成的指南路徑
        """
        try:
            logger.info("開始生成AI腳本開發指南")
            
            if not project_data:
                project_data = self._load_default_project_data()
            
            # 生成欄位使用指南
            field_guide = self._generate_field_usage_guide(project_data)
            
            # 生成邏輯模式指南
            logic_guide = self._generate_logic_pattern_guide(project_data)
            
            # 生成最佳實踐指南
            best_practices = self._generate_best_practices_guide(project_data)
            
            # 整合所有指南
            complete_guide = {
                "title": "Centimani AI腳本開發完整指南",
                "description": "基於實際腳本分析的開發指南和最佳實踐",
                "created_date": datetime.datetime.now().isoformat(),
                "field_usage_guide": field_guide,
                "logic_pattern_guide": logic_guide,
                "best_practices": best_practices,
                "examples": self._generate_examples(project_data)
            }
            
            # 儲存指南
            if not output_path:
                output_path = self.output_dir / f"ai_development_guide_{self.timestamp}.json"
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(complete_guide, f, ensure_ascii=False, indent=2)
            
            logger.info(f"AI開發指南生成完成: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"生成AI開發指南時發生錯誤: {e}")
            raise
    
    def generate_ini_files(self, 
                          project_name: str = "TestScript",
                          station_name: str = "Main",
                          output_dir: str = None) -> Dict[str, str]:
        """
        生成INI配置檔案
        
        Args:
            project_name: 專案名稱
            station_name: 站別名稱
            output_dir: 輸出目錄
            
        Returns:
            生成的INI檔案路徑字典
        """
        try:
            logger.info("開始生成INI配置檔案")
            
            if not output_dir:
                output_dir = self.output_dir / project_name
            
            output_dir = Path(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            
            ini_files = {}
            
            # 生成Global.ini
            global_ini_path = output_dir / "Global.ini"
            self._create_global_ini(global_ini_path, project_name)
            ini_files["global"] = str(global_ini_path)
            
            # 生成System.ini
            system_ini_path = output_dir / "System.ini"
            self._create_system_ini(system_ini_path, project_name, station_name)
            ini_files["system"] = str(system_ini_path)
            
            logger.info(f"INI檔案生成完成: {output_dir}")
            return ini_files
            
        except Exception as e:
            logger.error(f"生成INI檔案時發生錯誤: {e}")
            raise
    
    def validate_script(self, script_path: str) -> Dict[str, Any]:
        """
        驗證腳本的正確性
        
        Args:
            script_path: 腳本路徑
            
        Returns:
            驗證結果
        """
        try:
            logger.info(f"開始驗證腳本: {script_path}")
            
            script_path = Path(script_path)
            if not script_path.exists():
                raise FileNotFoundError(f"腳本不存在: {script_path}")
            
            wb = load_workbook(script_path, data_only=True)
            
            validation_result = {
                "script_path": str(script_path),
                "validation_timestamp": self.timestamp,
                "is_valid": True,
                "errors": [],
                "warnings": [],
                "sheet_validation": {},
                "overall_score": 100
            }
            
            # 驗證各個工作表
            required_sheets = ["Properties", "Instrument", "Switch", "DUTs"]
            for sheet_name in required_sheets:
                if sheet_name in wb.sheetnames:
                    sheet_validation = self._validate_sheet(wb[sheet_name], sheet_name)
                    validation_result["sheet_validation"][sheet_name] = sheet_validation
                    
                    if not sheet_validation["is_valid"]:
                        validation_result["is_valid"] = False
                        validation_result["errors"].extend(sheet_validation["errors"])
                        validation_result["overall_score"] -= 20
                else:
                    validation_result["errors"].append(f"缺少必要工作表: {sheet_name}")
                    validation_result["overall_score"] -= 25
            
            # 驗證測試工作表
            test_sheets = [name for name in wb.sheetnames if name not in required_sheets]
            for test_sheet in test_sheets:
                sheet_validation = self._validate_test_sheet(wb[test_sheet], test_sheet)
                validation_result["sheet_validation"][test_sheet] = sheet_validation
                
                if not sheet_validation["is_valid"]:
                    validation_result["warnings"].extend(sheet_validation["warnings"])
                    validation_result["overall_score"] -= 5
            
            validation_result["overall_score"] = max(0, validation_result["overall_score"])
            
            logger.info(f"腳本驗證完成: {script_path}")
            return validation_result
            
        except Exception as e:
            logger.error(f"驗證腳本時發生錯誤: {e}")
            return {
                "script_path": str(script_path),
                "validation_timestamp": self.timestamp,
                "is_valid": False,
                "errors": [str(e)],
                "warnings": [],
                "sheet_validation": {},
                "overall_score": 0
            }
    
    def _create_properties_sheet(self, wb: Workbook, project_name: str, station_name: str) -> None:
        """創建Properties工作表"""
        ws = wb.active
        ws.title = 'Properties'
        
        # 添加屬性
        properties = [
            ['Key', 'Value'],
            ['Version', '1.0.0'],
            ['Project', project_name],
            ['Station', station_name],
            ['CreateDate', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['FailRetryCount', '0'],
            ['Author', 'AI Assistant'],
            ['Description', f'{project_name} 測試腳本']
        ]
        
        for row in properties:
            ws.append(row)
    
    def _create_instrument_sheet(self, wb: Workbook, station_name: str) -> None:
        """創建Instrument工作表"""
        ws = wb.create_sheet('Instrument')
        
        # 添加表頭
        headers = ['Mount', 'Name', 'Sub', 'ControlDLL', 'Settings', 'Functions', 'Version', 'Switch', 'Desc']
        ws.append(headers)
        
        # 添加預設儀器
        instruments = [
            [station_name, 'DMM', '', 'DMM.dll', 'ADDR=ASRL1', 'MeasureV', '1.0.0', '', 'Digital Multimeter'],
            [station_name, 'PowerSupply', '', 'Power.dll', 'ADDR=ASRL2', 'SetVoltage,GetVoltage', '1.0.0', '', 'Power Supply']
        ]
        
        for instrument in instruments:
            ws.append(instrument)
    
    def _create_switch_sheet(self, wb: Workbook, station_name: str) -> None:
        """創建Switch工作表"""
        ws = wb.create_sheet('Switch')
        
        # 添加表頭
        headers = ['Mount', 'Name', 'ControlDLL', 'Settings', 'Version', 'Loss', 'Desc']
        ws.append(headers)
        
        # 添加預設開關
        switches = [
            [station_name, 'SW1', 'RF_Switch.dll', 'PORTS=8', '1.0.0', '', 'RF Switch'],
            [station_name, 'SW2', 'GPIO_Switch.dll', 'PORTS=16', '1.0.0', '', 'GPIO Switch']
        ]
        
        for switch in switches:
            ws.append(switch)
    
    def _create_duts_sheet(self, wb: Workbook, station_name: str) -> None:
        """創建DUTs工作表"""
        ws = wb.create_sheet('DUTs')
        
        # 添加表頭
        headers = ['Load', 'DUT Name', 'Station', 'Parameter', 'SFIF_DeviceID', 'ScanBarcode', 'Connection1']
        ws.append(headers)
        
        # 添加預設DUT
        duts = [
            ['true', f'SampleDUT_{station_name}', f'{station_name},0', '', 'DUT01', 
             '{ISN,12,^[A-Z0-9]+$}', 'Uart,UART1,COM3,115200,8,1,None,3000,5000']
        ]
        
        for dut in duts:
            ws.append(dut)
    
    def _create_main_test_sheet(self, wb: Workbook, station_name: str, template_config: Dict[str, Any]) -> None:
        """創建主要測試工作表"""
        ws = wb.create_sheet(station_name)
        
        # 添加表頭
        headers = ['TestID', 'Index', 'Phase', 'Title', 'Marco', 'Instrument', 'Switch', 'Send', 'Reply', 
                  'EndPrompt', 'Ref', 'Validation', 'Prompt', 'Waiting', 'ExecMode', 'Description']
        ws.append(headers)
        
        # 根據模板類型添加測試項目
        if template_config["test_items"] >= 3:
            # 基本三測項
            test_items = [
                ['V0001', '', 1, '版本號碼檢查', '', '', '', 'UART1@ver', '@Version:', '', 'version', '', '', '3000', '0', '讀取並檢查版本字串'],
                ['V0002', '', 2, '電壓測量', '', 'DMM.MeasureV', '', ':DMM.MeasureV,"CH1","DC"', '', '', 'voltage', '(voltage,%f,4.75,5.25)', '', '3000', '0', '量測 5V 軌電壓'],
                ['V0003', '', 3, 'BT狀態檢查', '', '', '', 'UART1@bt status', '@BT:ON', '', '', '', '', '3000', '0', '檢查藍牙啟用狀態']
            ]
            
            for item in test_items:
                ws.append(item)
        
        if template_config["test_items"] >= 5:
            # 額外測試項目
            extra_items = [
                ['V0004', '', 4, '網路連線測試', '', '', '', 'LAN@ping 8.8.8.8', '@PING:', '', 'network', '', '', '5000', '0', '測試網路連線'],
                ['V0005', '', 5, '音訊測試', '', 'Audio.dll', '', ':Audio.TestTone,1000,1000', '@AUDIO:', '', 'audio', '', '', '3000', '0', '測試音訊輸出']
            ]
            
            for item in extra_items:
                ws.append(item)
    
    def _create_config_sheet(self, wb: Workbook, station_name: str) -> None:
        """創建配置工作表"""
        ws = wb.create_sheet('Config')
        
        # 添加配置項目
        configs = [
            ['ConfigKey', 'ConfigValue', 'Description'],
            ['Timeout', '30000', '預設超時時間(ms)'],
            ['RetryCount', '3', '重試次數'],
            ['LogLevel', 'INFO', '日誌等級'],
            ['StationMode', 'AUTO', '站別模式']
        ]
        
        for config in configs:
            ws.append(config)
    
    def _create_macro_sheet(self, wb: Workbook, station_name: str) -> None:
        """創建宏工作表"""
        ws = wb.create_sheet('Macro')
        
        # 添加宏定義
        macros = [
            ['MacroName', 'MacroType', 'Parameters', 'Description'],
            ['PowerOn', 'Function', 'Delay:2000', '上電序列'],
            ['PowerOff', 'Function', 'Delay:1000', '下電序列'],
            ['Calibration', 'Function', 'Timeout:10000', '校準序列']
        ]
        
        for macro in macros:
            ws.append(macro)
    
    def _find_target_sheet(self, wb: Workbook, station_name: str) -> Optional[str]:
        """找到目標工作表"""
        # 優先使用指定的站別名稱
        if station_name in wb.sheetnames:
            return station_name
        
        # 尋找可能的測試工作表
        test_sheet_patterns = ['Main', 'Test', 'Function', 'MB']
        for pattern in test_sheet_patterns:
            for sheet_name in wb.sheetnames:
                if pattern.lower() in sheet_name.lower():
                    return sheet_name
        
        # 如果都找不到，返回第一個非標準工作表
        standard_sheets = {'Properties', 'Instrument', 'Switch', 'DUTs', 'Config'}
        for sheet_name in wb.sheetnames:
            if sheet_name not in standard_sheets:
                return sheet_name
        
        return None
    
    def _build_header_map(self, ws) -> Dict[str, int]:
        """建立表頭對應"""
        header_map = {}
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(1, col).value
            if header_value:
                header_map[str(header_value).strip()] = col
        return header_map
    
    def _clear_rows_keep_header(self, ws) -> None:
        """清除行數據，保留表頭"""
        for row in range(ws.max_row, 1, -1):
            ws.delete_rows(row)
    
    def _apply_modifications(self, ws, header_map: Dict[str, int], modifications: Dict[str, Any]) -> None:
        """應用修改配置"""
        for modification in modifications.get("test_items", []):
            self._append_test_item(ws, header_map, modification)
    
    def _apply_default_test_items(self, ws, header_map: Dict[str, int]) -> None:
        """應用預設測試項目"""
        default_items = [
            {
                "TestID": "V0001",
                "Phase": 1,
                "Title": "版本號碼檢查",
                "Send": "UART1@ver",
                "Reply": "@Version:",
                "Ref": "version",
                "Waiting": "3000",
                "ExecMode": "0",
                "Description": "讀取並檢查版本字串"
            },
            {
                "TestID": "V0002",
                "Phase": 2,
                "Title": "電壓測量",
                "Instrument": "DMM.MeasureV",
                "Send": ":DMM.MeasureV,\"CH1\",\"DC\"",
                "Ref": "voltage",
                "Validation": "(voltage,%f,4.75,5.25)",
                "Waiting": "3000",
                "ExecMode": "0",
                "Description": "量測 5V 軌電壓"
            },
            {
                "TestID": "V0003",
                "Phase": 3,
                "Title": "BT狀態檢查",
                "Send": "UART1@bt status",
                "Reply": "@BT:ON",
                "Waiting": "3000",
                "ExecMode": "0",
                "Description": "檢查藍牙啟用狀態"
            }
        ]
        
        for item in default_items:
            self._append_test_item(ws, header_map, item)
    
    def _append_test_item(self, ws, header_map: Dict[str, int], item: Dict[str, Any]) -> None:
        """添加測試項目行"""
        row_data = [''] * ws.max_column
        
        for field, value in item.items():
            if field in header_map:
                col = header_map[field]
                row_data[col - 1] = value
        
        ws.append(row_data)
    
    def _create_global_ini(self, file_path: Path, project_name: str) -> None:
        """創建Global.ini"""
        content = f"""[DEFAULT]
Company=Demo
Project={project_name}
Version=1.0.0
Description=Generated by Unified Script Generator
CreatedDate={datetime.datetime.now().strftime('%Y-%m-%d')}
"""
        file_path.write_text(content, encoding='utf-8')
    
    def _create_system_ini(self, file_path: Path, project_name: str, station_name: str) -> None:
        """創建System.ini"""
        content = f"""[DEFAULT]
LargeWindow=Default
RetryCount=3
DebugMode=0
DeviceSemaphore=Default
IndividualScanBarCode=Default
NewFormatScriptFile={project_name}_{station_name}.xlsx
TestItemCodeFile=Test Item Code V2.00_20241106.xlsx
StationList={station_name}
StationName={station_name}
"""
        file_path.write_text(content, encoding='utf-8')
    
    def _load_default_project_data(self) -> Dict[str, Any]:
        """載入預設專案資料"""
        return {
            "default": {
                "field_definitions": {
                    "field_categories": {
                        "phase_fields": ["Phase", "Index", "TestID"],
                        "action_fields": ["Send", "Instrument", "Switch"],
                        "response_fields": ["Reply", "EndPrompt"],
                        "validation_fields": ["Validation", "Ref"],
                        "control_fields": ["ExecMode", "Waiting", "Prompt"]
                    }
                }
            }
        }
    
    def _generate_field_usage_guide(self, project_data: Dict[str, Any]) -> Dict[str, Any]:
        """生成欄位使用指南"""
        return {
            "title": "欄位使用指南",
            "description": "基於實際腳本分析的欄位使用模式",
            "field_categories": project_data.get("default", {}).get("field_definitions", {}).get("field_categories", {}),
            "usage_patterns": {
                "phase_based_testing": "使用Phase欄位控制測試順序",
                "instrument_control": "使用Instrument和Switch欄位控制設備",
                "data_validation": "使用Validation和Ref欄位驗證結果"
            }
        }
    
    def _generate_logic_pattern_guide(self, project_data: Dict[str, Any]) -> Dict[str, Any]:
        """生成邏輯模式指南"""
        return {
            "title": "邏輯模式指南",
            "description": "常見的腳本邏輯模式",
            "patterns": {
                "sequential_testing": "順序執行測試項目",
                "conditional_testing": "根據條件執行不同測試路徑",
                "error_handling": "錯誤處理和重試機制",
                "data_collection": "數據收集和驗證"
            }
        }
    
    def _generate_best_practices_guide(self, project_data: Dict[str, Any]) -> Dict[str, Any]:
        """生成最佳實踐指南"""
        return {
            "title": "最佳實踐指南",
            "description": "腳本開發的最佳實踐",
            "practices": [
                "使用清晰的TestID和Title",
                "合理設定Phase值",
                "提供適當的錯誤處理",
                "使用有意義的變數名稱",
                "保持腳本結構一致"
            ]
        }
    
    def _generate_examples(self, project_data: Dict[str, Any]) -> Dict[str, Any]:
        """生成範例"""
        return {
            "basic_test_item": {
                "TestID": "V0001",
                "Phase": 1,
                "Title": "版本號碼檢查",
                "Send": "UART1@ver",
                "Reply": "@Version:",
                "Description": "讀取並檢查版本字串"
            },
            "instrument_test": {
                "TestID": "V0002",
                "Phase": 2,
                "Title": "電壓測量",
                "Instrument": "DMM.MeasureV",
                "Send": ":DMM.MeasureV,\"CH1\",\"DC\"",
                "Validation": "(voltage,%f,4.75,5.25)",
                "Description": "量測電壓並驗證範圍"
            }
        }
    
    def _validate_sheet(self, ws, sheet_name: str) -> Dict[str, Any]:
        """驗證工作表"""
        validation = {
            "sheet_name": sheet_name,
            "is_valid": True,
            "errors": [],
            "warnings": []
        }
        
        # 檢查是否有數據
        if ws.max_row <= 1:
            validation["warnings"].append("工作表沒有數據行")
        
        # 檢查表頭
        if ws.max_column < 2:
            validation["errors"].append("工作表列數不足")
            validation["is_valid"] = False
        
        return validation
    
    def _validate_test_sheet(self, ws, sheet_name: str) -> Dict[str, Any]:
        """驗證測試工作表"""
        validation = {
            "sheet_name": sheet_name,
            "is_valid": True,
            "errors": [],
            "warnings": []
        }
        
        # 檢查必要欄位
        required_headers = ["TestID", "Phase", "Title", "Send"]
        headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            validation["warnings"].append(f"缺少建議欄位: {', '.join(missing_headers)}")
        
        # 檢查測試項目
        if ws.max_row <= 1:
            validation["warnings"].append("沒有測試項目")
        
        return validation

def main():
    """主函數"""
    parser = argparse.ArgumentParser(description='統一的Centimani腳本生成器')
    parser.add_argument('--mode', '-m', required=True, 
                       choices=['new', 'modify', 'guide', 'ini', 'validate'],
                       help='操作模式')
    
    # 新腳本生成參數
    parser.add_argument('--template', '-t', choices=['basic', 'standard', 'advanced'],
                       help='腳本模板類型')
    parser.add_argument('--station', '-s', default='Main', help='站別名稱')
    parser.add_argument('--project', '-p', default='TestScript', help='專案名稱')
    parser.add_argument('--output', '-o', help='輸出路徑')
    
    # 模板修改參數
    parser.add_argument('--source', help='源模板路徑')
    
    # 驗證參數
    parser.add_argument('--script', help='要驗證的腳本路徑')
    
    args = parser.parse_args()
    
    # 初始化生成器
    generator = UnifiedCentimaniScriptGenerator()
    
    try:
        if args.mode == 'new':
            # 生成新腳本
            if not args.template:
                args.template = 'basic'
            
            script_path = generator.generate_new_script(
                template_type=args.template,
                station_name=args.station,
                project_name=args.project,
                output_path=args.output
            )
            print(f"新腳本已生成: {script_path}")
            
        elif args.mode == 'modify':
            # 從模板修改
            if not args.source:
                print("錯誤: 修改模式需要指定源模板路徑 (--source)")
                return 1
            
            script_path = generator.modify_from_template(
                template_path=args.source,
                output_path=args.output or f"modified_{args.station}_{generator.timestamp}.xlsx",
                station_name=args.station
            )
            print(f"腳本修改完成: {script_path}")
            
        elif args.mode == 'guide':
            # 生成開發指南
            guide_path = generator.generate_development_guide(output_path=args.output)
            print(f"開發指南已生成: {guide_path}")
            
        elif args.mode == 'ini':
            # 生成INI檔案
            ini_files = generator.generate_ini_files(
                project_name=args.project,
                station_name=args.station,
                output_dir=args.output
            )
            print("INI檔案已生成:")
            for file_type, file_path in ini_files.items():
                print(f"  {file_type}: {file_path}")
            
        elif args.mode == 'validate':
            # 驗證腳本
            if not args.script:
                print("錯誤: 驗證模式需要指定腳本路徑 (--script)")
                return 1
            
            validation_result = generator.validate_script(args.script)
            print(f"腳本驗證完成:")
            print(f"  有效性: {'是' if validation_result['is_valid'] else '否'}")
            print(f"  分數: {validation_result['overall_score']}/100")
            
            if validation_result['errors']:
                print("  錯誤:")
                for error in validation_result['errors']:
                    print(f"    - {error}")
            
            if validation_result['warnings']:
                print("  警告:")
                for warning in validation_result['warnings']:
                    print(f"    - {warning}")
        
        print("操作完成！")
        return 0
        
    except Exception as e:
        print(f"操作過程中發生錯誤: {e}")
        return 1

if __name__ == "__main__":
    exit(main()) 