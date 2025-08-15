#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
統一的Centimani腳本分析器
Unified Centimani Script Analyzer

整合了三個現有工具的所有功能：
1. universal_centimani_script_analyzer.py - 通用腳本分析
2. centimani_analysis_tool.py - CENTIMANIA分析
3. analyze_scripts_logic.py - 腳本邏輯分析

主要功能：
1. 自動識別腳本類型和專案
2. 分析腳本結構和組織方式
3. 識別測試流程和邏輯
4. 分析目錄結構和Excel檔案
5. 生成標準化的分析報告
6. 支援多種輸出格式
7. 整合的腳本邏輯分析

作者: AI Assistant
版本: 2.0.0
創建日期: 2025-01-XX
更新日期: 2025-01-XX
"""

import pandas as pd
import json
import os
import re
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('unified_script_analysis.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class UnifiedCentimaniScriptAnalyzer:
    """統一的Centimani腳本分析器"""
    
    def __init__(self, base_path: str = None):
        """
        初始化分析器
        
        Args:
            base_path: 基礎路徑，預設為當前目錄
        """
        self.base_path = Path(base_path) if base_path else Path.cwd()
        self.script_path = ""
        self.analysis_results = {}
        self.script_type = ""
        self.project_name = ""
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 建立輸出目錄
        self.output_dir = self.base_path / "unified_analysis_output"
        self.output_dir.mkdir(exist_ok=True)
        
        # 預定義的腳本類型識別規則
        self.script_type_patterns = {
            'kangaroo': {
                'keywords': ['kangaroo', 'kang', 'kangaroo_'],
                'file_patterns': [r'kangaroo', r'kang'],
                'sheet_patterns': [r'mb', r'function', r'dut', r'instrument']
            },
            'valo360': {
                'keywords': ['valo360', 'valo', 'valo360_'],
                'file_patterns': [r'valo360', r'valo'],
                'sheet_patterns': [r'mb', r'temp', r'function', r'poe']
            },
            'raptor': {
                'keywords': ['raptor', 'raptor_'],
                'file_patterns': [r'raptor'],
                'sheet_patterns': [r'test', r'function', r'fixture']
            },
            'kilimanjaro': {
                'keywords': ['kilimanjaro', 'kili'],
                'file_patterns': [r'kilimanjaro', r'kili'],
                'sheet_patterns': [r'led', r'function']
            },
            'zebra': {
                'keywords': ['zebra', 'zebra_'],
                'file_patterns': [r'zebra'],
                'sheet_patterns': [r'wifi', r'function']
            }
        }
        
        # 通用工作表類型識別
        self.sheet_type_patterns = {
            'properties': [r'properties', r'prop', r'config'],
            'instrument': [r'instrument', r'inst', r'device'],
            'switch': [r'switch', r'sw', r'relay'],
            'duts': [r'dut', r'device', r'equipment'],
            'test': [r'test', r'function', r'mb', r'phase'],
            'config': [r'config', r'setting', r'param']
        }
        
        # 腳本邏輯分析模式
        self.cli_pattern = re.compile(r"^[^:]+@\S+")  # e.g., UART1@ver
        self.device_call_pattern = re.compile(r"^:\S+")  # e.g., :DMM.MeasureV
        self.known_meta_sheets = {"Properties", "Instrument", "Switch", "DUTs"}
        
        logger.info(f"初始化統一的Centimani腳本分析器")
        logger.info(f"基礎路徑: {self.base_path}")
        logger.info(f"輸出目錄: {self.output_dir}")
    
    def analyze_script(self, script_path: str) -> Dict[str, Any]:
        """
        分析單一腳本文件
        
        Args:
            script_path: 腳本文件路徑
            
        Returns:
            腳本分析結果
        """
        try:
            self.script_path = script_path
            logger.info(f"開始分析腳本: {script_path}")
            
            # 檢查文件是否存在
            if not os.path.exists(script_path):
                raise FileNotFoundError(f"腳本文件不存在: {script_path}")
            
            # 識別腳本類型和專案
            self._identify_script_type(script_path)
            
            # 分析腳本結構
            structure_analysis = self._analyze_script_structure(script_path)
            
            # 分析腳本邏輯
            logic_analysis = self._analyze_script_logic(script_path)
            
            # 整合分析結果
            self.analysis_results = {
                "script_path": script_path,
                "script_type": self.script_type,
                "project_name": self.project_name,
                "analysis_timestamp": self.timestamp,
                "structure_analysis": structure_analysis,
                "logic_analysis": logic_analysis,
                "summary": self._generate_summary(structure_analysis, logic_analysis)
            }
            
            logger.info(f"腳本分析完成: {script_path}")
            return self.analysis_results
            
        except Exception as e:
            logger.error(f"分析腳本時發生錯誤: {e}")
            raise
    
    def analyze_directory_structure(self, target_dir: str) -> Dict[str, Any]:
        """
        分析目錄結構
        
        Args:
            target_dir: 目標目錄路徑
            
        Returns:
            目錄結構分析結果
        """
        logger.info(f"開始分析目錄結構: {target_dir}")
        
        target_path = Path(target_dir)
        if not target_path.exists():
            logger.error(f"目標目錄不存在: {target_dir}")
            return {}
        
        structure_analysis = {
            "directory_path": str(target_path),
            "total_files": 0,
            "file_types": {},
            "total_size": 0,
            "subdirectories": [],
            "files": [],
            "centimani_files": []
        }
        
        try:
            for item in target_path.rglob("*"):
                if item.is_file():
                    structure_analysis["total_files"] += 1
                    structure_analysis["total_size"] += item.stat().st_size
                    
                    # 統計文件類型
                    file_ext = item.suffix.lower()
                    if file_ext:
                        structure_analysis["file_types"][file_ext] = \
                            structure_analysis["file_types"].get(file_ext, 0) + 1
                    
                    # 識別Centimani相關文件
                    if self._is_centimani_file(item):
                        structure_analysis["centimani_files"].append({
                            "name": item.name,
                            "path": str(item.relative_to(target_path)),
                            "size": item.stat().st_size,
                            "type": self._identify_file_type(item)
                        })
                    
                    # 記錄文件信息
                    file_info = {
                        "name": item.name,
                        "path": str(item.relative_to(target_path)),
                        "size": item.stat().st_size,
                        "modified": datetime.fromtimestamp(item.stat().st_mtime).isoformat()
                    }
                    structure_analysis["files"].append(file_info)
                    
                elif item.is_dir():
                    structure_analysis["subdirectories"].append(str(item.relative_to(target_path)))
            
            logger.info(f"目錄結構分析完成: {target_dir}")
            return structure_analysis
            
        except Exception as e:
            logger.error(f"分析目錄結構時發生錯誤: {e}")
            return {}
    
    def analyze_excel_files(self, excel_files: List[str]) -> Dict[str, Any]:
        """
        分析Excel文件
        
        Args:
            excel_files: Excel文件路徑列表
            
        Returns:
            Excel文件分析結果
        """
        logger.info(f"開始分析Excel文件: {len(excel_files)}個文件")
        
        excel_analysis = {
            "total_files": len(excel_files),
            "files": [],
            "summary": {
                "total_sheets": 0,
                "total_rows": 0,
                "file_types": Counter()
            }
        }
        
        try:
            for excel_file in excel_files:
                if os.path.exists(excel_file):
                    file_analysis = self._analyze_single_excel(excel_file)
                    excel_analysis["files"].append(file_analysis)
                    
                    # 更新統計
                    excel_analysis["summary"]["total_sheets"] += len(file_analysis.get("sheets", []))
                    excel_analysis["summary"]["total_rows"] += file_analysis.get("total_rows", 0)
                    
                    # 統計文件類型
                    file_type = file_analysis.get("file_type", "unknown")
                    excel_analysis["summary"]["file_types"][file_type] += 1
            
            logger.info(f"Excel文件分析完成: {len(excel_files)}個文件")
            return excel_analysis
            
        except Exception as e:
            logger.error(f"分析Excel文件時發生錯誤: {e}")
            return {}
    
    def _identify_script_type(self, script_path: str) -> None:
        """識別腳本類型和專案"""
        filename = os.path.basename(script_path).lower()
        
        for script_type, patterns in self.script_type_patterns.items():
            # 檢查文件名關鍵字
            if any(keyword in filename for keyword in patterns['keywords']):
                self.script_type = script_type
                self.project_name = script_type.upper()
                logger.info(f"識別腳本類型: {self.script_type}")
                return
            
            # 檢查文件名模式
            if any(re.search(pattern, filename) for pattern in patterns['file_patterns']):
                self.script_type = script_type
                self.project_name = script_type.upper()
                logger.info(f"識別腳本類型: {self.script_type}")
                return
        
        # 如果無法識別，設為通用類型
        self.script_type = "generic"
        self.project_name = "UNKNOWN"
        logger.warning(f"無法識別腳本類型，設為通用類型")
    
    def _analyze_script_structure(self, script_path: str) -> Dict[str, Any]:
        """分析腳本結構"""
        try:
            wb = load_workbook(script_path, data_only=True)
            
            structure_analysis = {
                "file_path": script_path,
                "total_sheets": len(wb.sheetnames),
                "sheets": [],
                "sheet_types": Counter(),
                "total_rows": 0,
                "total_columns": 0
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_analysis = self._analyze_sheet_structure(ws, sheet_name)
                structure_analysis["sheets"].append(sheet_analysis)
                
                # 更新統計
                structure_analysis["total_rows"] += sheet_analysis.get("rows", 0)
                structure_analysis["total_columns"] = max(
                    structure_analysis["total_columns"], 
                    sheet_analysis.get("columns", 0)
                )
                
                # 統計工作表類型
                sheet_type = sheet_analysis.get("type", "unknown")
                structure_analysis["sheet_types"][sheet_type] += 1
            
            return structure_analysis
            
        except Exception as e:
            logger.error(f"分析腳本結構時發生錯誤: {e}")
            return {}
    
    def _analyze_sheet_structure(self, ws, sheet_name: str) -> Dict[str, Any]:
        """分析工作表結構"""
        try:
            # 讀取表頭
            headers = []
            for c in range(1, ws.max_column + 1):
                val = ws.cell(1, c).value
                headers.append(str(val).strip() if val is not None else "")
            
            # 識別工作表類型
            sheet_type = self._identify_sheet_type(sheet_name, headers)
            
            # 分析數據行
            data_rows = 0
            for r in range(2, ws.max_row + 1):
                if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
                    data_rows += 1
            
            return {
                "name": sheet_name,
                "type": sheet_type,
                "headers": headers,
                "rows": data_rows,
                "columns": ws.max_column,
                "has_data": data_rows > 0
            }
            
        except Exception as e:
            logger.error(f"分析工作表結構時發生錯誤: {e}")
            return {"name": sheet_name, "error": str(e)}
    
    def _identify_sheet_type(self, sheet_name: str, headers: List[str]) -> str:
        """識別工作表類型"""
        sheet_name_lower = sheet_name.lower()
        headers_lower = [h.lower() for h in headers]
        
        for sheet_type, patterns in self.sheet_type_patterns.items():
            # 檢查工作表名稱
            if any(re.search(pattern, sheet_name_lower) for pattern in patterns):
                return sheet_type
            
            # 檢查表頭關鍵字
            if any(any(re.search(pattern, h) for pattern in patterns) for h in headers_lower):
                return sheet_type
        
        return "unknown"
    
    def _analyze_script_logic(self, script_path: str) -> Dict[str, Any]:
        """分析腳本邏輯"""
        try:
            wb = load_workbook(script_path, data_only=True)
            
            logic_analysis = {
                "file_path": script_path,
                "test_flows": [],
                "statistics": {
                    "total_commands": 0,
                    "cli_commands": 0,
                    "device_calls": 0,
                    "phases": Counter(),
                    "exec_modes": Counter()
                }
            }
            
            for sheet_name in wb.sheetnames:
                if sheet_name not in self.known_meta_sheets:
                    # 分析測試工作表
                    ws = wb[sheet_name]
                    test_flow = self._analyze_test_flow(ws, sheet_name)
                    if test_flow:
                        logic_analysis["test_flows"].append(test_flow)
                        
                        # 更新統計
                        logic_analysis["statistics"]["total_commands"] += test_flow.get("total_commands", 0)
                        logic_analysis["statistics"]["cli_commands"] += test_flow.get("cli_commands", 0)
                        logic_analysis["statistics"]["device_calls"] += test_flow.get("device_calls", 0)
                        
                        # 統計Phase和ExecMode
                        for phase in test_flow.get("phases", []):
                            logic_analysis["statistics"]["phases"][phase] += 1
                        for exec_mode in test_flow.get("exec_modes", []):
                            logic_analysis["statistics"]["exec_modes"][exec_mode] += 1
            
            return logic_analysis
            
        except Exception as e:
            logger.error(f"分析腳本邏輯時發生錯誤: {e}")
            return {}
    
    def _analyze_test_flow(self, ws, sheet_name: str) -> Optional[Dict[str, Any]]:
        """分析測試流程"""
        try:
            headers = read_headers(ws)
            idx = {h: i + 1 for i, h in enumerate(headers) if h}
            
            # 檢查是否有必要的欄位
            if not any(col in idx for col in ["Send", "send", "Phase", "phase"]):
                return None
            
            test_flow = {
                "sheet_name": sheet_name,
                "total_commands": 0,
                "cli_commands": 0,
                "device_calls": 0,
                "phases": [],
                "exec_modes": [],
                "commands": []
            }
            
            for r in range(2, ws.max_row + 1):
                command_info = self._extract_command_info(ws, r, idx)
                if command_info:
                    test_flow["commands"].append(command_info)
                    test_flow["total_commands"] += 1
                    
                    # 統計命令類型
                    if command_info.get("is_cli"):
                        test_flow["cli_commands"] += 1
                    if command_info.get("is_device_call"):
                        test_flow["device_calls"] += 1
                    
                    # 記錄Phase和ExecMode
                    if command_info.get("phase"):
                        test_flow["phases"].append(command_info["phase"])
                    if command_info.get("exec_mode"):
                        test_flow["exec_modes"].append(command_info["exec_mode"])
            
            return test_flow if test_flow["total_commands"] > 0 else None
            
        except Exception as e:
            logger.error(f"分析測試流程時發生錯誤: {e}")
            return None
    
    def _extract_command_info(self, ws, row: int, idx: Dict[str, int]) -> Optional[Dict[str, Any]]:
        """提取命令信息"""
        try:
            # 獲取Send欄位
            send_col = idx.get("Send") or idx.get("send")
            if not send_col:
                return None
            
            send_value = ws.cell(row, send_col).value
            if not send_value:
                return None
            
            send_str = str(send_value).strip()
            if not send_str:
                return None
            
            command_info = {
                "row": row,
                "send": send_str,
                "is_cli": bool(self.cli_pattern.match(send_str)),
                "is_device_call": bool(self.device_call_pattern.match(send_str))
            }
            
            # 獲取Phase欄位
            phase_col = idx.get("Phase") or idx.get("phase")
            if phase_col:
                phase_value = ws.cell(row, phase_col).value
                if phase_value is not None:
                    command_info["phase"] = str(phase_value).strip()
            
            # 獲取ExecMode欄位
            execmode_col = idx.get("ExecMode")
            if execmode_col:
                execmode_value = ws.cell(row, execmode_col).value
                if execmode_value is not None:
                    command_info["exec_mode"] = str(execmode_value).strip()
            
            return command_info
            
        except Exception as e:
            logger.error(f"提取命令信息時發生錯誤: {e}")
            return None
    
    def _is_centimani_file(self, file_path: Path) -> bool:
        """判斷是否為Centimani相關文件"""
        file_name = file_path.name.lower()
        centimani_patterns = [
            r'centimani', r'kangaroo', r'valo360', r'raptor', 
            r'kilimanjaro', r'zebra', r'testflow', r'script'
        ]
        return any(re.search(pattern, file_name) for pattern in centimani_patterns)
    
    def _identify_file_type(self, file_path: Path) -> str:
        """識別文件類型"""
        file_name = file_path.name.lower()
        if file_name.endswith('.xlsx'):
            return "excel_script"
        elif file_name.endswith('.ini'):
            return "config_file"
        elif file_name.endswith('.log'):
            return "log_file"
        elif file_name.endswith('.md'):
            return "documentation"
        elif file_name.endswith('.json'):
            return "data_file"
        else:
            return "other"
    
    def _analyze_single_excel(self, excel_file: str) -> Dict[str, Any]:
        """分析單一Excel文件"""
        try:
            wb = load_workbook(excel_file, data_only=True)
            
            excel_analysis = {
                "file_path": excel_file,
                "file_name": os.path.basename(excel_file),
                "total_sheets": len(wb.sheetnames),
                "sheets": [],
                "total_rows": 0,
                "file_type": "unknown"
            }
            
            # 識別文件類型
            excel_analysis["file_type"] = self._identify_centimani_file_type(excel_file)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_analysis = self._analyze_sheet_structure(ws, sheet_name)
                excel_analysis["sheets"].append(sheet_analysis)
                excel_analysis["total_rows"] += sheet_analysis.get("rows", 0)
            
            return excel_analysis
            
        except Exception as e:
            logger.error(f"分析Excel文件時發生錯誤: {e}")
            return {"file_path": excel_file, "error": str(e)}
    
    def _identify_centimani_file_type(self, file_path: str) -> str:
        """識別Centimani文件類型"""
        file_name = os.path.basename(file_path).lower()
        
        if 'testflow' in file_name:
            return "test_flow"
        elif 'instrument' in file_name:
            return "instrument_config"
        elif 'ini' in file_path:
            return "config_file"
        elif 'script' in file_name:
            return "script_file"
        else:
            return "data_file"
    
    def _generate_summary(self, structure_analysis: Dict, logic_analysis: Dict) -> Dict[str, Any]:
        """生成分析摘要"""
        return {
            "script_type": self.script_type,
            "project_name": self.project_name,
            "total_sheets": structure_analysis.get("total_sheets", 0),
            "total_commands": logic_analysis.get("statistics", {}).get("total_commands", 0),
            "cli_ratio": self._calculate_ratio(
                logic_analysis.get("statistics", {}).get("cli_commands", 0),
                logic_analysis.get("statistics", {}).get("total_commands", 0)
            ),
            "device_call_ratio": self._calculate_ratio(
                logic_analysis.get("statistics", {}).get("device_calls", 0),
                logic_analysis.get("statistics", {}).get("total_commands", 0)
            ),
            "complexity_score": self._calculate_complexity_score(structure_analysis, logic_analysis)
        }
    
    def _calculate_ratio(self, numerator: int, denominator: int) -> float:
        """計算比例"""
        if denominator == 0:
            return 0.0
        return round(numerator / denominator * 100, 2)
    
    def _calculate_complexity_score(self, structure_analysis: Dict, logic_analysis: Dict) -> int:
        """計算複雜度分數"""
        score = 0
        
        # 工作表數量
        score += min(structure_analysis.get("total_sheets", 0) * 2, 20)
        
        # 命令數量
        score += min(logic_analysis.get("statistics", {}).get("total_commands", 0) // 10, 30)
        
        # 工作表類型多樣性
        sheet_types = structure_analysis.get("sheet_types", {})
        score += min(len(sheet_types) * 5, 25)
        
        # Phase數量
        phases = logic_analysis.get("statistics", {}).get("phases", {})
        score += min(len(phases) * 3, 25)
        
        return min(score, 100)
    
    def save_analysis_report(self, output_path: str = None) -> str:
        """保存分析報告"""
        if not output_path:
            output_path = self.output_dir / f"unified_analysis_{self.timestamp}.json"
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.analysis_results, f, ensure_ascii=False, indent=2)
            
            logger.info(f"分析報告已保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"保存分析報告時發生錯誤: {e}")
            raise
    
    def generate_markdown_report(self, output_path: str = None) -> str:
        """生成Markdown報告"""
        if not output_path:
            output_path = self.output_dir / f"unified_analysis_{self.timestamp}.md"
        
        try:
            markdown_content = self._generate_markdown_content()
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
            
            logger.info(f"Markdown報告已生成: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"生成Markdown報告時發生錯誤: {e}")
            raise
    
    def _generate_markdown_content(self) -> str:
        """生成Markdown內容"""
        if not self.analysis_results:
            return "# 無分析結果\n\n請先執行腳本分析。"
        
        content = f"""# Centimani腳本統一分析報告

## 基本資訊
- **腳本路徑**: {self.analysis_results.get('script_path', 'N/A')}
- **腳本類型**: {self.analysis_results.get('script_type', 'N/A')}
- **專案名稱**: {self.analysis_results.get('project_name', 'N/A')}
- **分析時間**: {self.analysis_results.get('analysis_timestamp', 'N/A')}

## 結構分析摘要
- **總工作表數**: {self.analysis_results.get('structure_analysis', {}).get('total_sheets', 0)}
- **總行數**: {self.analysis_results.get('structure_analysis', {}).get('total_rows', 0)}
- **工作表類型分布**: {dict(self.analysis_results.get('structure_analysis', {}).get('sheet_types', {}))}

## 邏輯分析摘要
- **總命令數**: {self.analysis_results.get('logic_analysis', {}).get('statistics', {}).get('total_commands', 0)}
- **CLI命令數**: {self.analysis_results.get('logic_analysis', {}).get('statistics', {}).get('cli_commands', 0)}
- **設備調用數**: {self.analysis_results.get('logic_analysis', {}).get('statistics', {}).get('device_calls', 0)}

## 綜合評估
- **複雜度分數**: {self.analysis_results.get('summary', {}).get('complexity_score', 0)}/100
- **CLI使用比例**: {self.analysis_results.get('summary', {}).get('cli_ratio', 0)}%
- **設備調用比例**: {self.analysis_results.get('summary', {}).get('device_call_ratio', 0)}%

## 詳細分析結果

### 工作表分析
"""
        
        # 添加工作表詳細信息
        for sheet in self.analysis_results.get('structure_analysis', {}).get('sheets', []):
            content += f"""
#### {sheet.get('name', 'Unknown')}
- **類型**: {sheet.get('type', 'Unknown')}
- **行數**: {sheet.get('rows', 0)}
- **列數**: {sheet.get('columns', 0)}
- **有數據**: {'是' if sheet.get('has_data', False) else '否'}
"""
        
        # 添加測試流程信息
        content += "\n### 測試流程分析\n"
        for flow in self.analysis_results.get('logic_analysis', {}).get('test_flows', []):
            content += f"""
#### {flow.get('sheet_name', 'Unknown')}
- **總命令數**: {flow.get('total_commands', 0)}
- **CLI命令**: {flow.get('cli_commands', 0)}
- **設備調用**: {flow.get('device_calls', 0)}
- **Phase數量**: {len(flow.get('phases', []))}
"""
        
        content += "\n---\n*本報告由統一的Centimani腳本分析器生成*"
        return content

def read_headers(ws):
    """讀取工作表表頭"""
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append(str(v).strip() if v is not None else "")
    return headers

def main():
    """主函數"""
    import argparse
    
    parser = argparse.ArgumentParser(description='統一的Centimani腳本分析器')
    parser.add_argument('--script', '-s', help='要分析的腳本文件路徑')
    parser.add_argument('--directory', '-d', help='要分析的目錄路徑')
    parser.add_argument('--excel', '-e', nargs='+', help='要分析的Excel文件路徑列表')
    parser.add_argument('--output', '-o', help='輸出目錄路徑')
    parser.add_argument('--format', '-f', choices=['json', 'md', 'both'], default='both', 
                       help='輸出格式 (json, md, both)')
    
    args = parser.parse_args()
    
    # 初始化分析器
    analyzer = UnifiedCentimaniScriptAnalyzer(args.output)
    
    try:
        if args.script:
            # 分析單一腳本
            results = analyzer.analyze_script(args.script)
            print(f"腳本分析完成: {args.script}")
            
        elif args.directory:
            # 分析目錄結構
            results = analyzer.analyze_directory_structure(args.directory)
            print(f"目錄分析完成: {args.directory}")
            
        elif args.excel:
            # 分析Excel文件
            results = analyzer.analyze_excel_files(args.excel)
            print(f"Excel文件分析完成: {len(args.excel)}個文件")
            
        else:
            print("請指定要分析的腳本、目錄或Excel文件")
            return
        
        # 保存結果
        if args.format in ['json', 'both']:
            json_path = analyzer.save_analysis_report()
            print(f"JSON報告已保存: {json_path}")
        
        if args.format in ['md', 'both']:
            md_path = analyzer.generate_markdown_report()
            print(f"Markdown報告已生成: {md_path}")
        
        print("分析完成！")
        
    except Exception as e:
        print(f"分析過程中發生錯誤: {e}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main()) 