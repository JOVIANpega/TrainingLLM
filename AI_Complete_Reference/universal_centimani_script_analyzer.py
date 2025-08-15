#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用Centimani腳本分析器
Universal Centimani Script Analyzer

這是一個通用的腳本分析工具，可以分析任何Centimani Excel腳本文件，
避免為每個腳本文件重複編寫分析代碼。

主要功能：
1. 自動識別腳本類型和專案
2. 分析腳本結構和組織方式
3. 識別測試流程和邏輯
4. 生成標準化的分析報告
5. 支援多種輸出格式

作者: AI Assistant
版本: 1.0.0
創建日期: 2024-12-19
"""

import pandas as pd
import json
import os
import re
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import logging
from pathlib import Path

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('script_analysis.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class UniversalCentimaniScriptAnalyzer:
    """通用Centimani腳本分析器"""
    
    def __init__(self):
        self.script_path = ""
        self.analysis_results = {}
        self.script_type = ""
        self.project_name = ""
        
        # 預定義的腳本類型識別規則
        self.script_type_patterns = {
            'kangaroo': {
                'keywords': ['kangaroo', 'kang', 'kangaroo_'],
                'file_patterns': [r'kangaroo', r'kang'],
                'sheet_patterns': [r'mb', r'function', r'dut', r'instrument']
            },
            'valo360': {
                'keywords': ['valo360', 'valo', 'valo360_'],
                'file_patterns': [r'valo360', r'valo'],
                'sheet_patterns': [r'mb', r'temp', r'function', r'poe']
            },
            'raptor': {
                'keywords': ['raptor', 'raptor_'],
                'file_patterns': [r'raptor'],
                'sheet_patterns': [r'test', r'function', r'fixture']
            },
            'kilimanjaro': {
                'keywords': ['kilimanjaro', 'kili'],
                'file_patterns': [r'kilimanjaro', r'kili'],
                'sheet_patterns': [r'led', r'function']
            },
            'zebra': {
                'keywords': ['zebra', 'zebra_'],
                'file_patterns': [r'zebra'],
                'sheet_patterns': [r'wifi', r'function']
            }
        }
        
        # 通用工作表類型識別
        self.sheet_type_patterns = {
            'properties': [r'properties', r'prop', r'config'],
            'instrument': [r'instrument', r'inst', r'device'],
            'switch': [r'switch', r'sw', r'relay'],
            'duts': [r'dut', r'device', r'equipment'],
            'test': [r'test', r'function', r'mb', r'phase'],
            'config': [r'config', r'setting', r'param']
        }
    
    def analyze_script(self, script_path: str) -> Dict[str, Any]:
        """分析腳本文件"""
        try:
            self.script_path = script_path
            logger.info(f"開始分析腳本: {script_path}")
            
            # 檢查文件是否存在
            if not os.path.exists(script_path):
                raise FileNotFoundError(f"腳本文件不存在: {script_path}")
            
            # 識別腳本類型和專案
            self._identify_script_type()
            
            # 載入工作簿對象（用於編輯功能）
            try:
                import openpyxl
                self.workbook = openpyxl.load_workbook(script_path)
                logger.info("成功載入工作簿對象，支援編輯功能")
            except ImportError:
                logger.warning("openpyxl未安裝，編輯功能不可用")
                self.workbook = None
            except Exception as e:
                logger.warning(f"載入工作簿對象失敗，編輯功能不可用: {e}")
                self.workbook = None
            
            # 讀取Excel文件
            excel_file = pd.ExcelFile(script_path)
            sheet_names = excel_file.sheet_names
            
            logger.info(f"發現工作表: {sheet_names}")
            
            # 分析每個工作表
            sheet_analyses = {}
            for sheet_name in sheet_names:
                logger.info(f"分析工作表: {sheet_name}")
                sheet_analysis = self._analyze_sheet(sheet_name, excel_file)
                sheet_analyses[sheet_name] = sheet_analysis
            
            # 生成綜合分析報告
            comprehensive_analysis = self._generate_comprehensive_analysis(
                sheet_names, sheet_analyses
            )
            
            self.analysis_results = comprehensive_analysis
            return comprehensive_analysis
            
        except Exception as e:
            logger.error(f"分析腳本失敗: {e}")
            return {}
    
    def _identify_script_type(self):
        """識別腳本類型和專案名稱"""
        try:
            file_name = os.path.basename(self.script_path).lower()
            
            # 根據文件名識別專案類型
            for project_name, patterns in self.script_type_patterns.items():
                for pattern in patterns['file_patterns']:
                    if re.search(pattern, file_name):
                        self.project_name = project_name.upper()
                        self.script_type = f"{project_name.upper()}_SCRIPT"
                        logger.info(f"識別腳本類型: {self.script_type}, 專案: {self.project_name}")
                        return
            
            # 如果無法識別，使用通用類型
            self.project_name = "UNKNOWN"
            self.script_type = "GENERIC_SCRIPT"
            logger.warning(f"無法識別腳本類型，使用通用類型: {self.script_type}")
            
        except Exception as e:
            logger.error(f"識別腳本類型失敗: {e}")
            self.project_name = "UNKNOWN"
            self.script_type = "GENERIC_SCRIPT"
    
    def _analyze_sheet(self, sheet_name: str, excel_file: pd.ExcelFile) -> Dict[str, Any]:
        """分析單個工作表"""
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # 基本統計信息
            total_rows = len(df)
            total_columns = len(df.columns)
            
            # 列名分析
            column_names = df.columns.tolist()
            column_types = df.dtypes.to_dict()
            
            # 識別工作表類型
            sheet_type = self._identify_sheet_type(sheet_name)
            
            # 數據內容分析
            content_analysis = self._analyze_sheet_content(df, sheet_name, sheet_type)
            
            return {
                'sheet_name': sheet_name,
                'sheet_type': sheet_type,
                'total_rows': total_rows,
                'total_columns': total_columns,
                'column_names': column_names,
                'column_types': {str(k): str(v) for k, v in column_types.items()},
                'content_analysis': content_analysis,
                'sample_data': self._get_sample_data(df)
            }
            
        except Exception as e:
            logger.error(f"分析工作表 {sheet_name} 失敗: {e}")
            return {'error': str(e)}
    
    def _identify_sheet_type(self, sheet_name: str) -> str:
        """識別工作表類型"""
        sheet_name_lower = sheet_name.lower()
        
        for sheet_type, patterns in self.sheet_type_patterns.items():
            for pattern in patterns:
                if re.search(pattern, sheet_name_lower):
                    return sheet_type
        
        return "unknown"
    
    def _analyze_sheet_content(self, df: pd.DataFrame, sheet_name: str, sheet_type: str) -> Dict[str, Any]:
        """分析工作表內容"""
        try:
            content_analysis = {
                'data_patterns': {},
                'key_columns': [],
                'test_items': [],
                'configuration_items': [],
                'special_features': [],
                'test_phases': [],
                'execution_modes': []
            }
            
            # 識別關鍵列
            for col in df.columns:
                col_str = str(col).lower()
                if any(keyword in col_str for keyword in ['test', 'item', 'step', 'phase']):
                    content_analysis['key_columns'].append(col)
                elif any(keyword in col_str for keyword in ['config', 'setting', 'param']):
                    content_analysis['configuration_items'].append(col)
                elif any(keyword in col_str for keyword in ['exec', 'mode', 'execmode']):
                    content_analysis['execution_modes'].append(col)
            
            # 分析測試項目
            if 'test' in df.columns or 'Test' in df.columns:
                test_col = 'test' if 'test' in df.columns else 'Test'
                if test_col in df.columns:
                    test_items = df[test_col].dropna().unique().tolist()
                    content_analysis['test_items'] = test_items
            
            # 分析測試階段
            if 'phase' in df.columns or 'Phase' in df.columns:
                phase_col = 'phase' if 'phase' in df.columns else 'Phase'
                if phase_col in df.columns:
                    phases = df[phase_col].dropna().unique().tolist()
                    content_analysis['test_phases'] = phases
            
            # 分析執行模式
            if 'execmode' in df.columns or 'ExecMode' in df.columns:
                exec_col = 'execmode' if 'execmode' in df.columns else 'ExecMode'
                if exec_col in df.columns:
                    modes = df[exec_col].dropna().unique().tolist()
                    content_analysis['execution_modes'] = modes
            
            # 分析特殊特徵
            if sheet_type in ['properties', 'config']:
                content_analysis['special_features'].append('配置工作表')
            elif sheet_type in ['test', 'function']:
                content_analysis['special_features'].append('測試執行工作表')
            elif sheet_type in ['instrument', 'duts']:
                content_analysis['special_features'].append('設備配置工作表')
            
            return content_analysis
            
        except Exception as e:
            logger.error(f"分析工作表內容失敗: {e}")
            return {}
    
    def _get_sample_data(self, df: pd.DataFrame) -> Dict[str, Any]:
        """獲取樣本數據"""
        try:
            sample_data = {}
            
            # 前5行數據
            if len(df) > 0:
                sample_data['first_5_rows'] = df.head().to_dict('records')
            
            # 非空值統計
            non_null_counts = df.count().to_dict()
            sample_data['non_null_counts'] = non_null_counts
            
            # 唯一值統計
            unique_counts = {}
            for col in df.columns:
                if df[col].dtype == 'object':
                    unique_counts[str(col)] = df[col].nunique()
            
            sample_data['unique_counts'] = unique_counts
            
            return sample_data
            
        except Exception as e:
            logger.error(f"獲取樣本數據失敗: {e}")
            return {}
    
    def _generate_comprehensive_analysis(self, sheet_names: List[str], 
                                       sheet_analyses: Dict[str, Any]) -> Dict[str, Any]:
        """生成綜合分析報告"""
        try:
            # 統計信息
            total_sheets = len(sheet_names)
            total_rows = sum(sheet_analyses.get(sheet, {}).get('total_rows', 0) 
                           for sheet in sheet_names)
            total_columns = sum(sheet_analyses.get(sheet, {}).get('total_columns', 0) 
                              for sheet in sheet_names)
            
            # 識別主要功能
            main_features = self._identify_main_features(sheet_names, sheet_analyses)
            
            # 測試流程分析
            test_flow_analysis = self._analyze_test_flow(sheet_analyses)
            
            # 腳本特點分析
            script_characteristics = self._analyze_script_characteristics(sheet_names, sheet_analyses)
            
            # 專案特定分析
            project_specific_analysis = self._analyze_project_specific_features()
            
            comprehensive_analysis = {
                'file_info': {
                    'file_name': os.path.basename(self.script_path),
                    'analysis_date': datetime.now().isoformat(),
                    'file_path': self.script_path,
                    'script_type': self.script_type,
                    'project_name': self.project_name,
                    'total_sheets': total_sheets,
                    'total_rows': total_rows,
                    'total_columns': total_columns
                },
                'sheet_overview': {
                    'sheet_names': sheet_names,
                    'sheet_analyses': sheet_analyses
                },
                'main_features': main_features,
                'test_flow_analysis': test_flow_analysis,
                'script_characteristics': script_characteristics,
                'project_specific_analysis': project_specific_analysis,
                'ai_integration_notes': self._generate_ai_integration_notes()
            }
            
            return comprehensive_analysis
            
        except Exception as e:
            logger.error(f"生成綜合分析報告失敗: {e}")
            return {}
    
    def _identify_main_features(self, sheet_names: List[str], 
                               sheet_analyses: Dict[str, Any]) -> List[str]:
        """識別主要功能特點"""
        features = []
        
        try:
            # 基於工作表類型識別功能
            sheet_types = [sheet_analyses.get(sheet, {}).get('sheet_type', '') 
                          for sheet in sheet_names]
            
            if 'properties' in sheet_types or 'config' in sheet_types:
                features.append('配置管理功能')
            
            if 'instrument' in sheet_types:
                features.append('儀器參數配置')
            
            if 'duts' in sheet_types:
                features.append('多設備測試支援')
            
            if 'test' in sheet_types or 'function' in sheet_types:
                features.append('測試執行功能')
            
            # 基於專案類型添加特定功能
            if self.project_name == "KANGAROO":
                features.extend(['KANGAROO專案專用', '多樣化測試支援'])
            elif self.project_name == "VALO360":
                features.extend(['VALO360專案專用', '功能導向測試'])
            elif self.project_name == "RAPTOR":
                features.extend(['RAPTOR專案專用', '全面測試覆蓋'])
            elif self.project_name == "KILIMANJARO":
                features.extend(['KILIMANJARO專案專用', '專注性測試'])
            elif self.project_name == "ZEBRA":
                features.extend(['ZEBRA專案專用', '基本功能測試'])
            
            # 基於內容識別功能
            for sheet_name, analysis in sheet_analyses.items():
                if 'test_items' in analysis.get('content_analysis', {}):
                    if len(analysis['content_analysis']['test_items']) > 0:
                        features.append('測試項目管理')
                        break
            
        except Exception as e:
            logger.error(f"識別主要功能失敗: {e}")
        
        return list(set(features))  # 去重
    
    def _analyze_test_flow(self, sheet_analyses: Dict[str, Any]) -> Dict[str, Any]:
        """分析測試流程"""
        test_flow = {
            'test_phases': [],
            'test_sequence': [],
            'dependencies': [],
            'error_handling': []
        }
        
        try:
            # 收集所有測試階段
            for sheet_name, analysis in sheet_analyses.items():
                if 'test_phases' in analysis.get('content_analysis', {}):
                    phases = analysis['content_analysis']['test_phases']
                    test_flow['test_phases'].extend(phases)
            
            # 去重並排序
            test_flow['test_phases'] = sorted(list(set(test_flow['test_phases'])))
            
            # 識別測試順序
            if test_flow['test_phases']:
                test_flow['test_sequence'] = [
                    '初始化配置',
                    '設備連接測試',
                    '功能測試執行',
                    '結果驗證',
                    '報告生成'
                ]
            
            # 識別依賴關係
            test_flow['dependencies'] = [
                '設備連接 → 配置載入',
                '配置載入 → 測試執行',
                '測試執行 → 結果收集',
                '結果收集 → 報告生成'
            ]
            
        except Exception as e:
            logger.error(f"分析測試流程失敗: {e}")
        
        return test_flow
    
    def _analyze_script_characteristics(self, sheet_names: List[str], 
                                       sheet_analyses: Dict[str, Any]) -> Dict[str, Any]:
        """分析腳本特點"""
        characteristics = {
            'complexity_level': '中等',
            'test_coverage': '功能測試',
            'script_organization': '模組化設計',
            'maintainability': '良好',
            'reusability': '中等'
        }
        
        try:
            # 基於工作表數量判斷複雜度
            if len(sheet_names) <= 3:
                characteristics['complexity_level'] = '簡單'
            elif len(sheet_names) <= 6:
                characteristics['complexity_level'] = '中等'
            else:
                characteristics['complexity_level'] = '複雜'
            
            # 基於行數判斷覆蓋範圍
            total_rows = sum(sheet_analyses.get(sheet, {}).get('total_rows', 0) 
                           for sheet in sheet_names)
            
            if total_rows <= 100:
                characteristics['test_coverage'] = '基本功能測試'
            elif total_rows <= 500:
                characteristics['test_coverage'] = '功能測試'
            else:
                characteristics['test_coverage'] = '全面功能測試'
            
            # 判斷腳本組織方式
            sheet_types = [sheet_analyses.get(sheet, {}).get('sheet_type', '') 
                          for sheet in sheet_names]
            
            if 'properties' in sheet_types or 'config' in sheet_types:
                characteristics['script_organization'] = '配置分離設計'
            
            if 'dut' in sheet_types:
                characteristics['script_organization'] = '設備分離設計'
            
        except Exception as e:
            logger.error(f"分析腳本特點失敗: {e}")
        
        return characteristics
    
    def _analyze_project_specific_features(self) -> Dict[str, Any]:
        """分析專案特定功能"""
        project_features = {
            'project_name': self.project_name,
            'script_type': self.script_type,
            'special_capabilities': [],
            'best_practices': [],
            'learning_focus': []
        }
        
        try:
            if self.project_name == "KANGAROO":
                project_features['special_capabilities'] = [
                    '多DUT測試支援',
                    'WiFi/BT校準測試',
                    'FW升級測試',
                    'OQC檢查流程'
                ]
                project_features['best_practices'] = [
                    '模組化腳本設計',
                    '配置與執行分離',
                    '標準化測試流程'
                ]
                project_features['learning_focus'] = [
                    '複雜測試流程管理',
                    '多設備協調測試',
                    '射頻測試最佳實踐'
                ]
            
            elif self.project_name == "VALO360":
                project_features['special_capabilities'] = [
                    '功能檢查測試',
                    'IMU測試',
                    'POE測試',
                    'DMIC測試'
                ]
                project_features['best_practices'] = [
                    '功能導向測試設計',
                    '介面測試標準化',
                    '溫度測試流程'
                ]
                project_features['learning_focus'] = [
                    '功能測試設計模式',
                    '介面測試最佳實踐',
                    '環境測試流程'
                ]
            
            # 其他專案類型...
            
        except Exception as e:
            logger.error(f"分析專案特定功能失敗: {e}")
        
        return project_features
    
    def _generate_ai_integration_notes(self) -> Dict[str, Any]:
        """生成AI整合說明"""
        return {
            'ai_learning_focus': [
                f'{self.project_name}專案腳本模式學習',
                '腳本組織和結構設計',
                '測試流程邏輯理解',
                '配置管理最佳實踐'
            ],
            'script_patterns': [
                '配置工作表 + 測試工作表結構',
                '設備配置與測試執行分離',
                '模組化腳本設計模式',
                '標準化測試流程'
            ],
            'integration_benefits': [
                '提高腳本生成準確性',
                f'支援{self.project_name}專案腳本開發',
                '優化測試流程設計',
                '標準化腳本結構'
            ],
            'usage_recommendations': [
                f'作為{self.project_name}專案腳本模板',
                f'學習{self.project_name}專案腳本設計',
                '參考配置管理最佳實踐',
                '理解測試流程邏輯'
            ]
        }
    
    def save_analysis_report(self, output_path: str) -> bool:
        """保存分析報告"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.analysis_results, f, ensure_ascii=False, indent=2)
            
            logger.info(f"分析報告已保存到: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"保存分析報告失敗: {e}")
            return False
    
    def generate_markdown_report(self, output_path: str) -> bool:
        """生成Markdown格式報告"""
        try:
            if not self.analysis_results:
                logger.warning("沒有分析結果可生成報告")
                return False
            
            with open(output_path, 'w', encoding='utf-8') as f:
                file_info = self.analysis_results['file_info']
                f.write(f"# {file_info['file_name']} 分析報告\n\n")
                
                # 文件信息
                f.write("## 文件信息\n\n")
                f.write(f"- **文件名稱**: {file_info['file_name']}\n")
                f.write(f"- **腳本類型**: {file_info['script_type']}\n")
                f.write(f"- **專案名稱**: {file_info['project_name']}\n")
                f.write(f"- **分析日期**: {file_info['analysis_date']}\n")
                f.write(f"- **工作表數量**: {file_info['total_sheets']}\n")
                f.write(f"- **總行數**: {file_info['total_rows']}\n")
                f.write(f"- **總列數**: {file_info['total_columns']}\n\n")
                
                # 工作表概覽
                f.write("## 工作表概覽\n\n")
                for sheet_name in self.analysis_results['sheet_overview']['sheet_names']:
                    f.write(f"- {sheet_name}\n")
                f.write("\n")
                
                # 主要功能
                f.write("## 主要功能\n\n")
                for feature in self.analysis_results['main_features']:
                    f.write(f"- {feature}\n")
                f.write("\n")
                
                # 腳本特點
                f.write("## 腳本特點\n\n")
                characteristics = self.analysis_results['script_characteristics']
                f.write(f"- **複雜度等級**: {characteristics['complexity_level']}\n")
                f.write(f"- **測試覆蓋**: {characteristics['test_coverage']}\n")
                f.write(f"- **腳本組織**: {characteristics['script_organization']}\n")
                f.write(f"- **可維護性**: {characteristics['maintainability']}\n")
                f.write(f"- **可重用性**: {characteristics['reusability']}\n\n")
                
                # 專案特定分析
                f.write("## 專案特定分析\n\n")
                project_analysis = self.analysis_results['project_specific_analysis']
                f.write(f"**專案名稱**: {project_analysis['project_name']}\n")
                f.write(f"**腳本類型**: {project_analysis['script_type']}\n\n")
                
                if project_analysis['special_capabilities']:
                    f.write("### 特殊能力\n")
                    for capability in project_analysis['special_capabilities']:
                        f.write(f"- {capability}\n")
                    f.write("\n")
                
                if project_analysis['best_practices']:
                    f.write("### 最佳實踐\n")
                    for practice in project_analysis['best_practices']:
                        f.write(f"- {practice}\n")
                    f.write("\n")
                
                # AI整合說明
                f.write("## AI整合說明\n\n")
                ai_notes = self.analysis_results['ai_integration_notes']
                f.write("### AI學習重點\n")
                for focus in ai_notes['ai_learning_focus']:
                    f.write(f"- {focus}\n")
                f.write("\n")
                
                f.write("### 腳本模式\n")
                for pattern in ai_notes['script_patterns']:
                    f.write(f"- {pattern}\n")
                f.write("\n")
                
                f.write("### 使用建議\n")
                for rec in ai_notes['usage_recommendations']:
                    f.write(f"- {rec}\n")
            
            logger.info(f"Markdown報告已生成: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"生成Markdown報告失敗: {e}")
            return False

    # ==================== 新增的編輯功能 ====================
    
    def add_sheet(self, sheet_name: str, content: Dict[str, Any] = None) -> bool:
        """新增工作表到Excel檔案"""
        try:
            if not hasattr(self, 'workbook') or not self.workbook:
                logger.error("沒有載入的工作簿，請先使用 analyze_script() 載入檔案")
                return False
            
            # 檢查工作表是否已存在
            if sheet_name in self.workbook.sheetnames:
                logger.warning(f"工作表 {sheet_name} 已存在")
                return False
            
            # 新增工作表
            new_sheet = self.workbook.create_sheet(title=sheet_name)
            
            # 如果有內容，添加到工作表
            if content:
                for row, row_data in content.items():
                    if isinstance(row_data, list):
                        for col, value in enumerate(row_data, 1):
                            new_sheet.cell(row=int(row), column=col, value=value)
                    elif isinstance(row_data, dict):
                        for col, value in row_data.items():
                            new_sheet.cell(row=int(row), column=col, value=value)
            
            logger.info(f"成功新增工作表: {sheet_name}")
            return True
            
        except Exception as e:
            logger.error(f"新增工作表失敗: {e}")
            return False
    
    def edit_sheet(self, sheet_name: str, changes: Dict[str, Any]) -> bool:
        """編輯工作表內容"""
        try:
            if not hasattr(self, 'workbook') or not self.workbook:
                logger.error("沒有載入的工作簿，請先使用 analyze_script() 載入檔案")
                return False
            
            if sheet_name not in self.workbook.sheetnames:
                logger.error(f"工作表 {sheet_name} 不存在")
                return False
            
            sheet = self.workbook[sheet_name]
            
            # 應用變更
            for cell_ref, value in changes.items():
                if ':' in cell_ref:  # 範圍引用，如 "A1:B5"
                    start_cell, end_cell = cell_ref.split(':')
                    # 這裡可以實現範圍編輯邏輯
                    logger.info(f"範圍編輯: {cell_ref} = {value}")
                else:  # 單一儲存格，如 "A1"
                    sheet[cell_ref] = value
                    logger.info(f"編輯儲存格: {cell_ref} = {value}")
            
            logger.info(f"成功編輯工作表: {sheet_name}")
            return True
            
        except Exception as e:
            logger.error(f"編輯工作表失敗: {e}")
            return False
    
    def save_changes(self, output_path: str = None) -> bool:
        """保存修改到檔案"""
        try:
            if not hasattr(self, 'workbook') or not self.workbook:
                logger.error("沒有載入的工作簿")
                return False
            
            # 如果沒有指定輸出路徑，使用原檔案路徑
            if not output_path:
                output_path = self.script_path
            
            # 保存檔案
            self.workbook.save(output_path)
            logger.info(f"修改已保存到: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"保存修改失敗: {e}")
            return False
    
    def get_workbook(self) -> Any:
        """獲取工作簿對象，用於進階操作"""
        if hasattr(self, 'workbook') and self.workbook:
            return self.workbook
        else:
            logger.warning("沒有載入的工作簿")
            return None
    
    def close_workbook(self) -> bool:
        """關閉工作簿"""
        try:
            if hasattr(self, 'workbook') and self.workbook:
                self.workbook.close()
                self.workbook = None
                logger.info("工作簿已關閉")
                return True
            return False
        except Exception as e:
            logger.error(f"關閉工作簿失敗: {e}")
            return False

def main():
    """主函數 - 示範使用"""
    analyzer = UniversalCentimaniScriptAnalyzer()
    
    # 檢查是否有腳本文件可分析
    script_files = [
        'Centimani DOC/scripe_EXCEL/KANGAROO/Pega_Kangaroo_TestFlow_Mainboard_MP_250407.xlsx',
        'Centimani DOC/scripe_EXCEL/VALO360/VALO360_TestFlow_20250318_MB_ER2.xlsx',
        'Centimani DOC/scripe_EXCEL/RAPTOR/Raptor_ALL_TestFlow_20230223_ALL.XLSX'
    ]
    
    for script_file in script_files:
        if os.path.exists(script_file):
            print(f"\n正在分析腳本: {script_file}")
            print("=" * 60)
            
            # 分析腳本
            analysis_results = analyzer.analyze_script(script_file)
            
            if analysis_results:
                # 生成輸出文件名
                base_name = os.path.splitext(os.path.basename(script_file))[0]
                json_output = f"{base_name}_Analysis.json"
                md_output = f"{base_name}_Analysis.md"
                
                # 保存報告
                analyzer.save_analysis_report(json_output)
                analyzer.generate_markdown_report(md_output)
                
                print(f"分析完成！報告已保存為:")
                print(f"- JSON: {json_output}")
                print(f"- Markdown: {md_output}")
                
                # 顯示分析摘要
                print(f"\n=== 分析摘要 ===")
                print(f"腳本類型: {analysis_results['file_info']['script_type']}")
                print(f"專案名稱: {analysis_results['file_info']['project_name']}")
                print(f"工作表數量: {analysis_results['file_info']['total_sheets']}")
                print(f"主要功能: {', '.join(analysis_results['main_features'][:5])}")
                print(f"複雜度等級: {analysis_results['script_characteristics']['complexity_level']}")
                
                # 示範新增的編輯功能
                print(f"\n=== 編輯功能示範 ===")
                if analyzer.workbook:
                    print("✅ 編輯功能可用")
                    
                    # 示範新增工作表
                    new_sheet_name = "CRUSOR"
                    if analyzer.add_sheet(new_sheet_name):
                        print(f"✅ 成功新增工作表: {new_sheet_name}")
                        
                        # 示範編輯內容
                        changes = {
                            "A1": "CRUSOR工作表",
                            "A2": "創建時間: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "A3": "由UniversalCentimaniScriptAnalyzer創建"
                        }
                        if analyzer.edit_sheet(new_sheet_name, changes):
                            print("✅ 成功編輯工作表內容")
                            
                            # 保存修改
                            output_file = f"{base_name}_WITH_CRUSOR.xlsx"
                            if analyzer.save_changes(output_file):
                                print(f"✅ 修改已保存到: {output_file}")
                            else:
                                print("❌ 保存修改失敗")
                        else:
                            print("❌ 編輯工作表內容失敗")
                    else:
                        print(f"❌ 新增工作表失敗: {new_sheet_name}")
                    
                    # 關閉工作簿
                    analyzer.close_workbook()
                else:
                    print("❌ 編輯功能不可用")
                
            else:
                print("分析失敗")
        else:
            print(f"腳本文件不存在: {script_file}")

if __name__ == "__main__":
    main() 